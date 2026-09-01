"""Phase 10C.2: client-only HTML, Markdown and Excel portfolio surfaces."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from pandas.testing import assert_frame_equal

from src.client_portfolio_view import build_client_portfolio_view
from src.excel_writer import (
    build_client_workbook,
    classifications_blocks,
    models_and_win_rates_blocks,
    portfolio_events_blocks,
    portfolio_stability_blocks,
)
from src.html_report import _jinja_env
from src.html_view_models import build_client_page_vm
from src.portfolio import ENGINE_OPTIMIZER
from src.portfolio_presentation import PORTFOLIO_UNAVAILABLE_NOTE
from src.report_writer import build_client_report
from tests.factories import build_synthetic_client_dataframe, build_synthetic_client_result
from tests.test_portfolio import _client_result, _portfolio_dataframe
from tests.test_portfolio_stability import _state_dataframe


def _available_result():
    return _client_result(_state_dataframe())


def _empty_result():
    dataframe = _portfolio_dataframe()
    for column in (
        "SCP_MODEL_OLDER_3M", "SCP_MODEL_RECENT_3M",
        "ML_BEST_MODEL_OLDER_3M", "ML_BEST_MODEL_RECENT_3M",
    ):
        dataframe[column] = None
    return _client_result(dataframe)


def _render_portfolio_component(view: dict, excel_url: str | None = "../client.xlsx") -> str:
    return _jinja_env().get_template("components/client_portfolio.html").render(
        page={"portfolio": view, "own_excel_url": excel_url},
    )


def _sheet_text(worksheet) -> str:
    return "\n".join(
        str(cell) for row in worksheet.iter_rows(values_only=True) for cell in row if cell is not None
    )


def _cell_below_header(worksheet, header: str):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == header:
                return worksheet.cell(row=cell.row + 1, column=cell.column)
    raise AssertionError(f"Header not found: {header}")


def test_client_view_available_content_is_compact_explicit_and_non_mutating():
    result = _available_result()
    source_models = result.portfolio.model_tables.by_engine_block_model
    snapshot = source_models.copy(deep=True)

    view = build_client_portfolio_view(result.portfolio)

    assert view["state"] == "AVAILABLE_WITH_CONTENT"
    assert len(view["coverage"]) == 4
    assert {group["engine"] for group in view["models"]} == {
        "SCP Classic Auto", "SCP Classic Optimizer",
    }
    assert all(len(group["rows"]) <= 8 for group in view["models"])
    assert any(row["sample_note"] == "Sin performance evaluable" for group in view["models"] for row in group["rows"])
    assert "causalidad" in view["stability_note"]
    assert view["performance_by_stability"]
    assert_frame_equal(source_models, snapshot)


def test_client_view_truncates_only_by_selection_frequency_and_reports_total():
    rows = []
    for index in range(10):
        row = _portfolio_dataframe().iloc[[0]].copy(deep=True)
        row["ID_CONFIGURATION"] = 5000 + index
        row["SCP_MODEL_OLDER_3M"] = f"Auto {index:02d}"
        row["SCP_MODEL_RECENT_3M"] = f"Auto {index:02d}"
        row["ML_BEST_MODEL_OLDER_3M"] = f"Optimizer {index:02d}"
        row["ML_BEST_MODEL_RECENT_3M"] = f"Optimizer {index:02d}"
        rows.append(row)
    result = _client_result(pd.concat(rows, ignore_index=True))
    view = build_client_portfolio_view(result.portfolio)

    assert all(group["total_rows"] == 10 for group in view["models"])
    assert all(group["truncated"] and len(group["rows"]) == 8 for group in view["models"])
    assert view["model_transitions"]["total_rows"] == 20
    assert len(view["model_transitions"]["rows"]) == 16
    assert {row["engine"] for row in view["model_transitions"]["rows"]} == {
        "SCP Classic Auto", "SCP Classic Optimizer",
    }


def test_client_view_distinguishes_available_empty_and_unavailable_without_legacy_fallback():
    empty = build_client_portfolio_view(_empty_result().portfolio)
    assert empty["state"] == "AVAILABLE_EMPTY"
    assert empty["coverage"]
    assert empty["models"] == []
    assert "no se observaron asignaciones" in empty["message"]

    historical = _client_result(build_synthetic_client_dataframe())
    historical.source.dataframe["ML_BEST_MODEL"] = "LegacyMustNotAppear"
    unavailable = build_client_portfolio_view(historical.portfolio)
    assert unavailable["state"] == "UNAVAILABLE"
    assert unavailable["message"] == PORTFOLIO_UNAVAILABLE_NOTE
    assert unavailable["coverage"] == []
    assert unavailable["models"] == []
    assert "LegacyMustNotAppear" not in str(unavailable)


def test_client_view_is_independent_from_forbidden_legacy_values_when_available():
    forbidden = (
        "SCP_BEST_MODEL", "ML_BEST_MODEL", "ML_CLASSIFICATION",
        "SCP_CLASSIFICATION", "SERIES_CLASSIFICATION", "ML_TYPE",
    )
    left = _state_dataframe()
    right = _state_dataframe()
    for index, column in enumerate(forbidden):
        left[column] = f"left-{index}"
        right[column] = f"right-{index}"
    assert build_client_portfolio_view(_client_result(left).portfolio) == build_client_portfolio_view(
        _client_result(right).portfolio
    )


def test_html_portfolio_available_empty_and_unavailable_states_and_excel_link():
    available_html = _render_portfolio_component(build_client_portfolio_view(_available_result().portfolio))
    assert "Selección observada y estabilidad por período" in available_html
    assert "SCP Classic Auto" in available_html
    assert "SCP Classic Optimizer" in available_html
    assert "Sin familia mapeada" in available_html
    assert "Performance descriptiva por estabilidad" in available_html
    assert "Excel individual" in available_html and "../client.xlsx" in available_html
    assert "clasificación × modelo" not in available_html.lower()

    empty_html = _render_portfolio_component(build_client_portfolio_view(_empty_result().portfolio))
    assert "no se observaron asignaciones" in empty_html
    assert "Cobertura de selección" in empty_html
    assert "Frecuencia de selección por modelo" not in empty_html

    unavailable_html = _render_portfolio_component(
        build_client_portfolio_view(_client_result(build_synthetic_client_dataframe()).portfolio),
    )
    assert PORTFOLIO_UNAVAILABLE_NOTE in unavailable_html
    assert "Metadata específica ausente" in unavailable_html
    assert "Cobertura de selección" not in unavailable_html


def test_html_portfolio_escapes_model_and_classification_values():
    dataframe = _state_dataframe()
    malicious = '<script>alert("portfolio")</script>&|'
    for column in (
        "SCP_MODEL_OLDER_3M", "SCP_MODEL_RECENT_3M",
        "ML_BEST_MODEL_OLDER_3M", "ML_BEST_MODEL_RECENT_3M",
        "ML_CLASSIFICATION_OLDER_3M", "ML_CLASSIFICATION_RECENT_3M",
    ):
        dataframe[column] = malicious
    html = _render_portfolio_component(build_client_portfolio_view(_client_result(dataframe).portfolio))
    assert '<script>alert("portfolio")' not in html
    assert "&lt;script&gt;alert" in html
    assert "&amp;" in html


def test_client_page_vm_exposes_portfolio_even_without_comparable_performance():
    result = build_synthetic_client_result(with_data=False)
    page = build_client_page_vm(result)
    assert page["kind"] == "no_performance"
    assert page["portfolio"]["state"] == "UNAVAILABLE"
    assert "ml_models" not in page and "scp_models" not in page and "classifications" not in page


def test_markdown_sections_10_to_12_are_summary_not_excel_dump():
    report = build_client_report(_available_result())
    section_10 = report.split("## 10.")[1].split("## 11.")[0]
    section_11 = report.split("## 11.")[1].split("## 12.")[0]
    section_12 = report.split("## 12.")[1].split("## 13.")[0]

    assert "Selección observada por modelo y período" in section_10
    assert "ordenado únicamente por frecuencia de selección" in section_10
    assert "08_models_and_win_rates" in section_10 and "17_portfolio_events" in section_10
    assert "Familias" in section_11 and "16_portfolio_stability" in section_11
    assert "Performance descriptiva por estabilidad" in section_11
    assert "Cobertura clasificación" in section_12 and "09_classifications" in section_12
    assert "clasificación × modelo" in section_12
    assert "metadata legacy" not in (section_10 + section_11 + section_12).lower()


def test_markdown_unavailable_preserves_other_historical_sections():
    report = build_client_report(build_synthetic_client_result(with_data=True))
    assert PORTFOLIO_UNAVAILABLE_NOTE in report
    assert "## 9. Impacto absoluto" in report
    assert "## 13. Exclusiones" in report
    assert "## 18. Diagnóstico Fase 8" in report


def test_markdown_escapes_portfolio_pipe_and_newline_values():
    dataframe = _state_dataframe()
    dataframe["ML_BEST_MODEL_OLDER_3M"] = "Modelo|con\nsalto"
    dataframe["ML_BEST_MODEL_RECENT_3M"] = "Modelo|con\nsalto"
    report = build_client_report(_client_result(dataframe))
    assert "Modelo\\|con salto" in report


def test_excel_block_builders_are_full_native_and_have_no_legacy_fallback():
    result = _available_result()
    model_blocks = models_and_win_rates_blocks(result)
    class_blocks = classifications_blocks(result)
    stability_blocks = portfolio_stability_blocks(result)
    event_blocks = portfolio_events_blocks(result)

    model_table = next(frame for title, frame in model_blocks if title == "Modelos - SCP Classic Optimizer")
    core_optimizer = result.portfolio.model_tables.by_engine_block_model
    assert len(model_table) == len(core_optimizer[core_optimizer["engine"] == ENGINE_OPTIMIZER])
    assert model_table["WAPE — SCP Classic Auto"].dtype.kind in "fc"
    assert model_table["Mejora de WAPE — SCP Classic Optimizer frente a SCP Classic Auto"].dtype.kind in "fc"
    assert any("Clasificación del Optimizer × modelo observado" == title for title, _ in class_blocks)
    assert any("Performance descriptiva por estabilidad" == title for title, _ in stability_blocks)
    assert any("Eventos canónicos auditables" in title for title, _ in event_blocks)
    assert "ML_BEST_MODEL" not in str(model_blocks + class_blocks + stability_blocks + event_blocks)


def test_excel_workbook_adds_16_17_without_renumbering_and_preserves_native_scales(tmp_path: Path):
    output = tmp_path / "portfolio.xlsx"
    result = _available_result()
    build_client_workbook(result, output)
    workbook = openpyxl.load_workbook(output)

    assert workbook.sheetnames[-4:] == [
        "14_pareto_absolute_impact", "15_phase8_bias_volume",
        "16_portfolio_stability", "17_portfolio_events",
    ]
    assert workbook.sheetnames[8:10] == ["08_models_and_win_rates", "09_classifications"]

    models = workbook["08_models_and_win_rates"]
    ratio_cell = _cell_below_header(models, "Cuota de selección sobre asignaciones posibles")
    improvement_cell = _cell_below_header(
        models, "Mejora de WAPE — SCP Classic Optimizer frente a SCP Classic Auto",
    )
    assert isinstance(ratio_cell.value, (int, float)) and 0 <= ratio_cell.value <= 1
    assert ratio_cell.number_format == "0.0%"
    assert isinstance(improvement_cell.value, (int, float)) and abs(improvement_cell.value) > 1
    assert improvement_cell.number_format == '+0.0"%";-0.0"%";0.0"%"'

    events = workbook["17_portfolio_events"]
    wape_cell = _cell_below_header(events, "WAPE — SCP Classic Auto")
    bias_cell = _cell_below_header(events, "Bias — SCP Classic Auto")
    assert wape_cell.value == 0.2 and wape_cell.number_format == "0.0%"
    assert bias_cell.value == 0.2 and bias_cell.number_format == "+0.0%;-0.0%;0.0%"
    assert events.max_row > 10


def test_excel_unavailable_keeps_all_portfolio_sheets_as_notes_only(tmp_path: Path):
    output = tmp_path / "unavailable.xlsx"
    build_client_workbook(build_synthetic_client_result(with_data=True), output)
    workbook = openpyxl.load_workbook(output)
    for name in (
        "08_models_and_win_rates", "09_classifications",
        "16_portfolio_stability", "17_portfolio_events",
    ):
        text = _sheet_text(workbook[name])
        assert "selección por bloques no disponible" in text
        assert "metadata legacy" not in text
        assert "AutoETS" not in text


def test_client_surfaces_never_recalculate_portfolio(monkeypatch, tmp_path: Path):
    result = _available_result()

    def _boom(*args, **kwargs):
        raise AssertionError("client surfaces must not recalculate portfolio analytics")

    for name in (
        "build_portfolio_analysis", "build_portfolio_event_table", "build_portfolio_coverage",
        "build_portfolio_model_result", "build_portfolio_optimizer_analysis",
        "build_portfolio_stability_analysis",
    ):
        if hasattr(__import__("src.portfolio", fromlist=[name]), name):
            monkeypatch.setattr(f"src.portfolio.{name}", _boom)

    before = result.portfolio.events.dataframe.copy(deep=True)
    build_client_portfolio_view(result.portfolio)
    build_client_page_vm(result)
    build_client_report(result)
    build_client_workbook(result, tmp_path / "no_recalc.xlsx")
    assert_frame_equal(result.portfolio.events.dataframe, before)
