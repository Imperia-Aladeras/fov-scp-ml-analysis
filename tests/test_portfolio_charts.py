"""Phase 10C.4: global portfolio charts and final output integration."""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from pandas.testing import assert_frame_equal

import analysis_fov_scp_ml as pipeline
from src.charts import generate_client_charts
from src.global_analysis import analyze_global
from src.global_charts import (
    PORTFOLIO_FAMILY_CHART,
    PORTFOLIO_STABILITY_CHART,
    _optimizer_family_selection_chart_data,
    _portfolio_stability_chart_data,
    generate_portfolio_charts,
)
from src.html_report import validate_run_links
from src.portfolio_presentation import (
    BLOCK_LABELS,
    COLUMN_PRESENTATIONS,
    PORTFOLIO_AVAILABLE_EMPTY_NOTE,
    PORTFOLIO_UNAVAILABLE_NOTE,
)
from src.report_writer import build_client_report
from src.global_report_writer import build_global_report
from tests.factories import build_global_analysis_result
from tests.test_global_portfolio_surfaces import (
    _available_global,
    _empty_client,
    _empty_global,
)
from tests.test_html_report import _build_client_row
from tests.test_portfolio import _client_result
from tests.test_portfolio_optimizer import _optimizer_dataframe


def _sheet_text(worksheet) -> str:
    return "\n".join(
        str(cell) for row in worksheet.iter_rows(values_only=True) for cell in row
        if cell is not None
    )


def _canonical_event_rows(worksheet) -> list[dict]:
    id_batch_label = COLUMN_PRESENTATIONS["ID_BATCH"].visible_label
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if id_batch_label not in row:
            continue
        headers = list(row)
        records = []
        for values in worksheet.iter_rows(min_row=row_index + 1, values_only=True):
            if not any(value is not None for value in values):
                break
            records.append(dict(zip(headers, values)))
        return records
    raise AssertionError("Canonical event header not found")


def _block_specific_dataframe() -> pd.DataFrame:
    rows = []
    block_values = (
        {
            "SCP_MODEL_OLDER_3M": "Auto estable",
            "SCP_MODEL_RECENT_3M": "Auto estable",
            "ML_BEST_MODEL_OLDER_3M": "AutoETS",
            "ML_BEST_MODEL_RECENT_3M": "AutoETS",
            "ML_CLASSIFICATION_OLDER_3M": "smooth",
            "ML_CLASSIFICATION_RECENT_3M": "smooth",
        },
        {
            "SCP_MODEL_OLDER_3M": "Auto anterior",
            "SCP_MODEL_RECENT_3M": "Auto reciente",
            "ML_BEST_MODEL_OLDER_3M": "FutureModel",
            "ML_BEST_MODEL_RECENT_3M": "LGBMRegressor",
            "ML_CLASSIFICATION_OLDER_3M": "erratic",
            "ML_CLASSIFICATION_RECENT_3M": "smooth",
        },
    )
    for index, values in enumerate(block_values, start=1):
        row = _build_client_row(
            10204,
            id_configuration=1000 + index,
            id_batch=63,
            id_run_staging=60,
            source_run_id=1,
            winner="ML" if index != 2 else "SCP",
            scp_err=20.0 if index != 2 else 10.0,
            ml_err=10.0 if index != 2 else 30.0,
        )
        row.update(values)
        rows.append(row)
    return pd.DataFrame(rows)


def test_generate_portfolio_charts_creates_only_two_deterministic_global_assets(tmp_path: Path):
    out_dir = tmp_path / "portfolio"
    generated = generate_portfolio_charts(_available_global(), out_dir)

    assert generated == [
        str(out_dir / PORTFOLIO_FAMILY_CHART),
        str(out_dir / PORTFOLIO_STABILITY_CHART),
    ]
    assert {path.name for path in out_dir.iterdir()} == {
        PORTFOLIO_FAMILY_CHART,
        PORTFOLIO_STABILITY_CHART,
    }
    assert all(Path(path).stat().st_size > 0 for path in generated)


def test_family_chart_uses_both_blocks_and_keeps_every_observed_family_including_unmapped():
    result = analyze_global([_client_result(_optimizer_dataframe())])
    source = result.portfolio.optimizer.family_tables
    data = _optimizer_family_selection_chart_data(result)

    assert data is not None
    assert data["blocks"] == (
        BLOCK_LABELS["OLDER_3M"],
        BLOCK_LABELS["RECENT_3M"],
    )
    assert set(data["families"]) == set(source["family"].map({
        "baselines": "Modelos base",
        "classical": "Modelos clásicos",
        "intermittent": "Demanda intermitente",
        "ml": "Aprendizaje automático",
        "UNMAPPED": "Sin familia mapeada",
    }))
    assert "Sin familia mapeada" in data["families"]
    assert all(sum(data["values"][block]) == pytest.approx(100) for block in data["blocks"])
    assert _optimizer_family_selection_chart_data(result) == data


def test_stability_chart_is_100_percent_over_evaluable_and_keeps_not_evaluable_as_context():
    result = _available_global()
    rows = _portfolio_stability_chart_data(result)

    assert rows is not None
    assert {row["engine"] for row in rows} == {
        "SCP Classic Auto", "SCP Classic Optimizer",
    }
    for row in rows:
        assert row["stable_pct"] + row["changed_pct"] == 100
        assert row["n_evaluable"] == row["stable_count"] + row["changed_count"]
        assert row["not_evaluable_count"] > 0
        assert row["n_evaluable"] + row["not_evaluable_count"] > row["n_evaluable"]


def test_stability_chart_copy_explicitly_keeps_not_evaluable_outside_denominator(monkeypatch, tmp_path: Path):
    captured: list[tuple[str, str]] = []

    def _capture_title(ax, title: str, subtitle: str) -> None:
        captured.append((title, subtitle))

    monkeypatch.setattr("src.global_charts._apply_title", _capture_title)
    assert len(generate_portfolio_charts(_available_global(), tmp_path / "portfolio")) == 2

    stability_title, stability_subtitle = captured[1]
    assert stability_title == "Estabilidad observada entre períodos"
    assert "No evaluables (fuera del denominador)" in stability_subtitle
    assert "SCP Classic Auto: 2" in stability_subtitle
    assert "SCP Classic Optimizer: 2" in stability_subtitle


def test_portfolio_charts_are_omitted_for_unavailable_and_available_empty(tmp_path: Path):
    unavailable_dir = tmp_path / "unavailable" / "portfolio"
    empty_dir = tmp_path / "empty" / "portfolio"

    assert generate_portfolio_charts(build_global_analysis_result(), unavailable_dir) == []
    assert generate_portfolio_charts(_empty_global(), empty_dir) == []
    assert not unavailable_dir.exists()
    assert not empty_dir.exists()
    assert PORTFOLIO_AVAILABLE_EMPTY_NOTE in build_client_report(_empty_client(97101))
    assert PORTFOLIO_AVAILABLE_EMPTY_NOTE in build_global_report(_empty_global())


def test_client_charts_never_create_portfolio_assets(tmp_path: Path):
    result = _available_global().client_results[0]
    generated = generate_client_charts(result, tmp_path / "charts")

    assert generated
    assert not any("portfolio" in Path(path).parts for path in generated)
    assert not (tmp_path / "charts" / "portfolio").exists()


def test_portfolio_charts_do_not_mutate_source_tables_or_call_analytics(monkeypatch, tmp_path: Path):
    result = _available_global()
    family_before = result.portfolio.optimizer.family_tables.copy(deep=True)
    stability_before = result.portfolio.stability.model_summary.copy(deep=True)

    def _boom(*args, **kwargs):
        raise AssertionError("portfolio charts must not recalculate analytics")

    for name in (
        "build_portfolio_analysis", "build_portfolio_event_table", "build_portfolio_coverage",
        "build_portfolio_model_result", "build_optimizer_portfolio_result",
        "build_portfolio_stability_analysis", "bias_aggregate", "period_wape_global",
    ):
        if hasattr(__import__("src.portfolio", fromlist=[name]), name):
            monkeypatch.setattr(f"src.portfolio.{name}", _boom)

    assert len(generate_portfolio_charts(result, tmp_path / "portfolio")) == 2
    assert_frame_equal(result.portfolio.optimizer.family_tables, family_before)
    assert_frame_equal(result.portfolio.stability.model_summary, stability_before)


def test_historical_pipeline_has_unavailable_portfolio_without_assets_links_or_legacy_fallback(tmp_path: Path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    pd.DataFrame([_build_client_row(10204)]).to_csv(
        data_dir / "TA_FOV_SCP_ML_historical.csv", index=False,
    )
    output_root = tmp_path / "runs"

    assert pipeline.main([
        "--input-dir", str(data_dir), "--output-root", str(output_root),
        "--run-name", "historical_10c4",
    ]) == 0
    run_dir = output_root / "historical_10c4"
    client_dir = run_dir / "clients" / "10204-sklum"
    global_html = (run_dir / "index.html").read_text(encoding="utf-8")
    client_html = (client_dir / "index.html").read_text(encoding="utf-8")
    global_portfolio = global_html.split('id="portfolio-global"')[1].split("</section>")[0]
    client_portfolio = client_html.split('id="seleccion-estabilidad"')[1].split("</section>")[0]
    manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))

    assert PORTFOLIO_UNAVAILABLE_NOTE in global_html and PORTFOLIO_UNAVAILABLE_NOTE in client_html
    assert PORTFOLIO_UNAVAILABLE_NOTE in (run_dir / "global" / "fov_scp_ml_global_report.md").read_text(encoding="utf-8")
    assert PORTFOLIO_UNAVAILABLE_NOTE in next(client_dir.glob("*.md")).read_text(encoding="utf-8")
    assert "AutoETS" not in global_portfolio and "AutoETS" not in client_portfolio
    assert not (run_dir / "global" / "charts" / "portfolio").exists()
    assert not any("charts/portfolio" in path for path in manifest["outputs_generated"])
    assert validate_run_links(run_dir) == []

    client_book = openpyxl.load_workbook(next(client_dir.glob("*.xlsx")), read_only=True)
    global_book = openpyxl.load_workbook(run_dir / "global" / "fov_scp_ml_global_summary.xlsx", read_only=True)
    assert "selección por bloques no disponible" in _sheet_text(client_book["17_portfolio_events"]).lower()
    assert "selección por bloques no disponible" in _sheet_text(global_book["19_portfolio_events"]).lower()


def test_block_specific_pipeline_integrates_global_portfolio_charts_outputs_and_traceability(tmp_path: Path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    _block_specific_dataframe().to_csv(
        data_dir / "TA_FOV_SCP_ML_block_specific.csv", index=False,
    )
    output_root = tmp_path / "runs"

    assert pipeline.main([
        "--input-dir", str(data_dir), "--output-root", str(output_root),
        "--run-name", "block_specific_10c4",
    ]) == 0
    run_dir = output_root / "block_specific_10c4"
    client_dir = run_dir / "clients" / "10204-sklum"
    portfolio_dir = run_dir / "global" / "charts" / "portfolio"
    expected_assets = {
        f"global/charts/portfolio/{PORTFOLIO_FAMILY_CHART}",
        f"global/charts/portfolio/{PORTFOLIO_STABILITY_CHART}",
    }
    manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
    outputs = {Path(path).as_posix() for path in manifest["outputs_generated"]}
    global_html = (run_dir / "index.html").read_text(encoding="utf-8")
    client_html = (client_dir / "index.html").read_text(encoding="utf-8")

    assert {path.name for path in portfolio_dir.iterdir()} == {
        PORTFOLIO_FAMILY_CHART, PORTFOLIO_STABILITY_CHART,
    }
    assert expected_assets <= outputs
    assert all(asset in global_html for asset in expected_assets)
    assert "charts/portfolio" not in client_html
    assert not (client_dir / "charts" / "portfolio").exists()
    assert "SCP Classic Auto" in global_html and "SCP Classic Optimizer" in global_html
    assert "3 meses anteriores (M6–M4)" in global_html
    assert "3 meses recientes (M3–M1)" in global_html
    assert "Sin familia mapeada" in global_html
    assert validate_run_links(run_dir) == []

    client_report = next(client_dir.glob("*.md")).read_text(encoding="utf-8")
    global_report = (run_dir / "global" / "fov_scp_ml_global_report.md").read_text(encoding="utf-8")
    assert "Selección observada" in client_report and "Selección observada" in global_report

    client_book = openpyxl.load_workbook(next(client_dir.glob("*.xlsx")), read_only=True)
    global_book = openpyxl.load_workbook(run_dir / "global" / "fov_scp_ml_global_summary.xlsx", read_only=True)
    client_events = _canonical_event_rows(client_book["17_portfolio_events"])
    global_events = _canonical_event_rows(global_book["19_portfolio_events"])
    assert len(client_events) == len(global_events) == 8
    identity_columns = [
        COLUMN_PRESENTATIONS[column].visible_label
        for column in (
            "ID_BATCH", "ID_RUN_STAGING", "ID_CLIENT", "SOURCE_RUN_ID",
            "ID_CONFIGURATION", "engine", "block",
        )
    ]
    assert all(all(row[column] is not None for column in identity_columns) for row in client_events)
    assert all(all(row[column] is not None for column in identity_columns) for row in global_events)

    moved = tmp_path / "moved" / run_dir.name
    moved.parent.mkdir()
    shutil.copytree(run_dir, moved)
    assert validate_run_links(moved) == []
