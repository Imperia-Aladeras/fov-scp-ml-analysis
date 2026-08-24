from pathlib import Path

import openpyxl

from src.client_analysis import analyze_client
from src.global_analysis import GlobalAnalysisResult, analyze_global
from src.global_excel_writer import (
    build_global_workbook,
    classifications_blocks,
    data_quality_checks_blocks,
    models_and_win_rates_blocks,
    pareto_absolute_impact_blocks,
    phase8_global_blocks,
    readme_blocks,
)
from src.quality_checks import QualityIssue, Severity
from tests.factories import (
    build_global_analysis_result,
    build_phase8_global_missing_client_results,
    build_phase8_global_multi_client_analysis_result,
    build_synthetic_client_dataframe,
    make_client_source,
)

EXPECTED_SHEETS = [
    "00_readme", "01_executive_summary", "02_client_coverage", "03_semester_by_client",
    "04_first_quarter_by_client", "05_second_quarter_by_client", "06_monthly_by_client",
    "07_global_period_summary", "08_client_improvement_stats", "09_series_improvement_stats",
    "10_winner_distribution", "11_models_and_win_rates", "12_classifications",
    "13_absolute_impact", "14_exclusions", "15_data_quality_checks", "16_pareto_absolute_impact",
    "17_phase8_global",
]


def test_build_global_workbook_creates_all_sheets(tmp_path: Path):
    result = build_global_analysis_result()
    out_path = tmp_path / "global_summary.xlsx"
    build_global_workbook(result, out_path)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_global_period_summary_has_one_row_per_period(tmp_path: Path):
    result = build_global_analysis_result()
    out_path = tmp_path / "global_summary.xlsx"
    build_global_workbook(result, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["07_global_period_summary"]
    # fila 1 = titulo del bloque, fila 2 = cabecera, filas 3+ = una por periodo
    period_cells = [r[0] for r in ws.iter_rows(min_row=3, values_only=True) if r[0]]
    assert len(period_cells) == 9


def test_readme_client_lists_show_id_client_alongside_display_name():
    """
    El listado de clientes del README global (00_readme) no debe mostrar
    unicamente el nombre de catalogo: el catalogo no garantiza nombres unicos
    entre ID_CLIENT distintos.
    """
    result = build_global_analysis_result()
    blocks = readme_blocks(result)
    lines = "\n".join(blocks[0][1][""].astype(str))

    assert "Synthetic (99999)" in lines
    assert "NoComparable (88888)" in lines
    assert "AllMlWins (77777)" in lines


def test_data_quality_checks_counts_affected_clients_by_id_not_by_shared_file_label(tmp_path: Path):
    """
    Fase 5, bug corregido: dos ClientAnalysisResult que comparten el mismo
    file_label (porque proceden del mismo CSV fisico multi-cliente, Fase 3)
    deben contarse como 2 clientes afectados, no colapsar en 1 por compartir
    file_label/display_name como clave de identidad.
    """
    df_a = build_synthetic_client_dataframe()
    source_a = make_client_source(df_a, 10001, "SharedLabel")
    result_a = analyze_client(source_a)

    df_b = build_synthetic_client_dataframe()
    df_b["ID_CLIENT"] = 10002
    source_b = make_client_source(df_b, 10002, "SharedLabel")
    # mismo CSV fisico: mismo file_label y display_name que source_a, pero ID_CLIENT distinto
    source_b.file_label = source_a.file_label
    source_b.display_name = source_a.display_name
    result_b = analyze_client(source_b)

    same_issue = QualityIssue(
        Severity.WARNING, "SHARED_TEST_ISSUE", "incidencia sintetica compartida", scope="file",
    )
    result_a.quality.add(same_issue)
    result_b.quality.add(same_issue)

    global_result = GlobalAnalysisResult(client_results=[result_a, result_b], invalid_results=[])
    blocks = data_quality_checks_blocks(global_result)
    summary_df = blocks[0][1]

    row = summary_df[summary_df["CODIGO"] == "SHARED_TEST_ISSUE"].iloc[0]
    assert row["N_OCURRENCIAS"] == 2
    assert row["N_CLIENTES_AFECTADOS"] == 2

    # El detalle fila a fila (misma hoja) tambien debe llevar ID_CLIENT: el
    # resumen ya deduplicaba bien, pero el detalle solo mostraba CLIENTE
    # (display_name), y ambos clientes sinteticos comparten display_name.
    detail_df = blocks[1][1]
    shared_detail = detail_df[detail_df["CODIGO"] == "SHARED_TEST_ISSUE"]
    assert set(shared_detail["ID_CLIENT"]) == {10001, 10002}
    assert list(shared_detail.columns[:2]) == ["ID_CLIENT", "CLIENTE"]


# --------------------------------------------------------------------------
# 16_pareto_absolute_impact
# --------------------------------------------------------------------------

def test_pareto_absolute_impact_blocks_contains_four_groups_and_summary():
    result = build_global_analysis_result()
    blocks = pareto_absolute_impact_blocks(result)
    titles = [t for t, _ in blocks]

    assert any("Pareto series - mejora" in t for t in titles)
    assert any("Pareto series - deterioro" in t for t in titles)
    assert any("Pareto clientes - mejora" in t for t in titles)
    assert any("Pareto clientes - deterioro" in t for t in titles)
    summary_title = next(t for t in titles if "Resumen de concentracion" in t)
    summary_table = next(df for t, df in blocks if t == summary_title)
    assert list(summary_table["GRUPO"]) == [
        "Series - mejora (ABS_ERROR_REDUCTION > 0)", "Series - deterioro (ABS_ERROR_REDUCTION < 0)",
        "Clientes - mejora (ABS_ERROR_REDUCTION > 0)", "Clientes - deterioro (ABS_ERROR_REDUCTION < 0)",
    ]


def test_pareto_absolute_impact_blocks_reads_precomputed_pareto_without_recomputing():
    """
    Igual que en la hoja individual 14: las tablas devueltas deben ser
    exactamente los mismos objetos DataFrame ya almacenados en
    GlobalPeriodResult.pareto_series/.pareto_clients (identidad de objeto,
    no solo igualdad de contenido).
    """
    result = build_global_analysis_result()
    gp = result.periods["6M"]
    blocks = {title: df for title, df in pareto_absolute_impact_blocks(result)}

    series_improve_title = next(t for t in blocks if "Pareto series - mejora" in t)
    series_deteriorate_title = next(t for t in blocks if "Pareto series - deterioro" in t)
    clients_improve_title = next(t for t in blocks if "Pareto clientes - mejora" in t)
    clients_deteriorate_title = next(t for t in blocks if "Pareto clientes - deterioro" in t)

    assert blocks[series_improve_title] is gp.pareto_series.improvement.table
    assert blocks[series_deteriorate_title] is gp.pareto_series.deterioration.table
    assert blocks[clients_improve_title] is gp.pareto_clients.improvement.table
    assert blocks[clients_deteriorate_title] is gp.pareto_clients.deterioration.table


def test_build_global_workbook_creates_sheet_16(tmp_path: Path):
    result = build_global_analysis_result()
    out_path = tmp_path / "global_summary.xlsx"
    build_global_workbook(result, out_path)

    wb = openpyxl.load_workbook(out_path)
    assert "16_pareto_absolute_impact" in wb.sheetnames
    assert wb.sheetnames == EXPECTED_SHEETS


# --------------------------------------------------------------------------
# 17_phase8_global (Fase 8D)
# --------------------------------------------------------------------------

def test_build_global_workbook_creates_sheet_17(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    out_path = tmp_path / "global_summary.xlsx"
    build_global_workbook(result, out_path)

    wb = openpyxl.load_workbook(out_path)
    assert "17_phase8_global" in wb.sheetnames
    assert wb.sheetnames == EXPECTED_SHEETS


def test_phase8_global_blocks_contains_bias_volume_cross_and_not_assignable_count():
    result = build_phase8_global_multi_client_analysis_result()
    blocks = phase8_global_blocks(result)
    titles = [t for t, _ in blocks]

    assert any("Bias agregado global" in t for t in titles)
    assert any("Volumen relativo global" in t for t in titles)
    assert any("SERIES_CLASSIFICATION x VOLUME_BUCKET" in t for t in titles)
    assert any("NOT_ASSIGNABLE" in t for t in titles)

    not_assignable_title = next(t for t in titles if "NOT_ASSIGNABLE" in t)
    not_assignable_table = next(df for t, df in blocks if t == not_assignable_title)
    assert not_assignable_table["N_CLIENTES_VOLUMEN_NOT_ASSIGNABLE"].iloc[0] == 1


def test_phase8_global_volume_block_never_has_n_clients_column():
    """
    Contrato del nucleo (src.global_analysis.build_phase8_global_diagnostics):
    volume_table global NUNCA lleva columna n_clients (a diferencia de
    model_tables/classification_tables/classification_volume_cross, que si
    la llevan). La hoja 17 no debe fabricarla.
    """
    result = build_phase8_global_multi_client_analysis_result()
    blocks = phase8_global_blocks(result)
    volume_title = next(t for t, _ in blocks if "Volumen relativo global" in t)
    volume_table = next(df for t, df in blocks if t == volume_title)
    assert "n_clients" not in {c.lower() for c in volume_table.columns}


def test_phase8_global_cross_table_includes_not_assignable_row():
    """El cruce global debe conservar filas SERIES_CLASSIFICATION x NOT_ASSIGNABLE tal cual las entrega el nucleo."""
    result = build_phase8_global_multi_client_analysis_result()
    blocks = phase8_global_blocks(result)
    cross_title = next(t for t, _ in blocks if "SERIES_CLASSIFICATION x VOLUME_BUCKET" in t)
    cross_table = next(df for t, df in blocks if t == cross_title)
    assert (cross_table["VOLUME_BUCKET"] == "No asignable").any()


def test_phase8_global_blocks_none_when_phase8_unavailable():
    result = analyze_global(build_phase8_global_missing_client_results())
    assert result.periods["6M"].phase8 is None
    blocks = phase8_global_blocks(result)
    assert len(blocks) == 1
    assert "no disponible" in blocks[0][1][""].iloc[0]


def test_models_and_classifications_blocks_include_bias_when_phase8_available():
    result = build_phase8_global_multi_client_analysis_result()
    ml_blocks = models_and_win_rates_blocks(result)
    ml_models_table = ml_blocks[0][1]
    assert "scp_bias_agg" in ml_models_table.columns
    assert "n_clients" in ml_models_table.columns

    class_blocks = classifications_blocks(result)
    series_class_table = next(df for t, df in class_blocks if t.startswith("SERIES_CLASSIFICATION"))
    assert "scp_bias_agg" in series_class_table.columns


def test_models_and_classifications_blocks_fall_back_without_bias_when_phase8_none():
    result = analyze_global(build_phase8_global_missing_client_results())
    assert result.periods["6M"].phase8 is None
    ml_blocks = models_and_win_rates_blocks(result)
    ml_models_table = ml_blocks[0][1]
    assert "scp_bias_agg" not in ml_models_table.columns
