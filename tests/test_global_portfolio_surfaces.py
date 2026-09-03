"""Phase 10C.3: pooled global portfolio HTML, Markdown, Excel and traceability."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from pandas.testing import assert_frame_equal

from src.global_analysis import analyze_global
from src.global_excel_writer import (
    build_global_workbook,
    classifications_blocks,
    models_and_win_rates_blocks,
    portfolio_events_blocks,
    portfolio_stability_blocks,
)
from src.global_portfolio_view import build_global_portfolio_view
from src.global_report_writer import build_global_report
from src.html_report import _jinja_env
from src.portfolio_presentation import PORTFOLIO_UNAVAILABLE_NOTE, prepare_portfolio_presentation
from tests.factories import build_global_analysis_result, build_synthetic_client_dataframe
from tests.test_portfolio import _client_result, _portfolio_dataframe
from tests.test_portfolio_stability import _state_dataframe


def _available_global():
    return analyze_global([
        _client_result(_state_dataframe(97001)),
        _client_result(_state_dataframe(97002)),
    ])


def _empty_client(id_client: int):
    dataframe = _portfolio_dataframe(id_client)
    for column in (
        "SCP_MODEL_OLDER_3M", "SCP_MODEL_RECENT_3M",
        "ML_BEST_MODEL_OLDER_3M", "ML_BEST_MODEL_RECENT_3M",
    ):
        dataframe[column] = None
    return _client_result(dataframe)


def _empty_global():
    return analyze_global([_empty_client(97101), _empty_client(97102)])


def _render_component(view: dict, excel_url: str | None = "global.xlsx") -> str:
    return _jinja_env().get_template("components/global_portfolio.html").render(
        global_portfolio=view,
        global_excel_url=excel_url,
    )


def _sheet_text(worksheet) -> str:
    return "\n".join(
        str(cell) for row in worksheet.iter_rows(values_only=True) for cell in row if cell is not None
    )


def _cell_below_header(worksheet, header: str):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == header:
                return worksheet.cell(row=cell.row + 1, column=cell.column)
    raise AssertionError(f"Header not found: {header}")


def test_global_view_uses_prepared_pooled_portfolio_with_clients_and_no_mutation():
    result = _available_global()
    source = result.portfolio.model_tables.by_engine_block_model
    snapshot = source.copy(deep=True)
    presentation = prepare_portfolio_presentation(result.portfolio)

    view = build_global_portfolio_view(presentation)

    assert view["state"] == "AVAILABLE_WITH_CONTENT"
    assert len(view["coverage"]) == 4
    assert {group["engine"] for group in view["models"]} == {
        "SCP Classic Auto", "SCP Classic Optimizer",
    }
    assert {row["n_clients"] for group in view["models"] for row in group["rows"]} == {"2", "0"}
    assert view["older_model_cohorts"]
    assert view["performance_by_stability"]
    assert all("n_clients" in row for row in view["performance_by_stability"])
    assert_frame_equal(source, snapshot)


def test_global_view_all_or_nothing_unavailable_and_available_empty():
    mixed = analyze_global([
        _client_result(_state_dataframe(97201)),
        _client_result(build_synthetic_client_dataframe(), id_client=97202),
    ])
    mixed_view = build_global_portfolio_view(prepare_portfolio_presentation(mixed.portfolio))
    assert mixed_view["state"] == "UNAVAILABLE"
    assert mixed_view["message"] == PORTFOLIO_UNAVAILABLE_NOTE
    assert mixed_view["coverage"] == []

    empty_view = build_global_portfolio_view(prepare_portfolio_presentation(_empty_global().portfolio))
    assert empty_view["state"] == "AVAILABLE_EMPTY"
    assert empty_view["coverage"]
    assert empty_view["models"] == []
    assert "no se observaron asignaciones" in empty_view["message"]


def test_global_view_reads_global_result_not_client_presentation_tables():
    result = _available_global()
    presentation = prepare_portfolio_presentation(result.portfolio)
    expected = build_global_portfolio_view(presentation)
    result.client_results.clear()
    assert build_global_portfolio_view(presentation) == expected


def test_global_html_contains_deeper_portfolio_content_and_methodology():
    view = build_global_portfolio_view(prepare_portfolio_presentation(_available_global().portfolio))
    html = _render_component(view)

    for text in (
        "Cobertura global de selección", "SCP Classic Auto", "SCP Classic Optimizer",
        "Familias del Optimizer", "Clasificación × modelo", "Clasificación × familia",
        "Cohortes por modelo del período anterior", "Transiciones de modelo más frecuentes",
        "Performance descriptiva por estabilidad", "N clientes", "Sin familia mapeada",
    ):
        assert text in html
    assert "pair-assignable" in html
    assert "no constituyen un ranking universal" in html.lower()
    assert "Excel global" in html


def test_global_html_unavailable_and_empty_do_not_render_fake_tables():
    unavailable = build_global_portfolio_view(
        prepare_portfolio_presentation(build_global_analysis_result().portfolio)
    )
    unavailable_html = _render_component(unavailable)
    assert PORTFOLIO_UNAVAILABLE_NOTE in unavailable_html
    assert "Cobertura global de selección" not in unavailable_html

    empty = build_global_portfolio_view(prepare_portfolio_presentation(_empty_global().portfolio))
    empty_html = _render_component(empty)
    assert "Sin asignaciones observadas" in empty_html
    assert "Cobertura global de selección" in empty_html
    assert "Selecciones de modelo más frecuentes" not in empty_html


def test_global_html_escapes_models_classifications_families_and_transitions():
    dataframe = _state_dataframe(97301)
    malicious = '<script>alert("global")</script>&'
    for column in (
        "SCP_MODEL_OLDER_3M", "SCP_MODEL_RECENT_3M",
        "ML_BEST_MODEL_OLDER_3M", "ML_BEST_MODEL_RECENT_3M",
        "ML_CLASSIFICATION_OLDER_3M", "ML_CLASSIFICATION_RECENT_3M",
    ):
        dataframe[column] = malicious
    result = analyze_global([_client_result(dataframe)])
    html = _render_component(build_global_portfolio_view(prepare_portfolio_presentation(result.portfolio)))
    assert '<script>alert("global")' not in html
    assert "&lt;script&gt;alert" in html
    assert "&amp;" in html


def test_global_markdown_sections_16_to_18_are_descriptive_and_reference_excel():
    report = build_global_report(_available_global())
    section_16 = report.split("## 16.")[1].split("## 17.")[0]
    section_17 = report.split("## 17.")[1].split("## 18.")[0]
    section_18 = report.split("## 18.")[1].split("## 19.")[0]

    assert "Selección observada" in section_16 and "N clientes" in section_16
    assert "11_models_and_win_rates" in section_16 and "19_portfolio_events" in section_16
    assert "Portfolio Optimizer" in section_17 and "pair-assignable" in section_17
    assert "12_classifications" in section_17
    assert "No evaluables" in section_18 and "Performance descriptiva" in section_18
    assert "18_portfolio_stability" in section_18
    combined = (section_16 + section_17 + section_18).lower()
    assert "metadata legacy" not in combined
    assert "familia recomendada" not in combined and "routing recomendado" not in combined


def test_global_markdown_unavailable_is_clear_and_keeps_other_sections():
    report = build_global_report(build_global_analysis_result())
    for start, end in (("## 16.", "## 17."), ("## 17.", "## 18."), ("## 18.", "## 19.")):
        assert PORTFOLIO_UNAVAILABLE_NOTE in report.split(start)[1].split(end)[0]
    assert "## 15." in report and "## 19." in report and "## 22." in report


def test_global_excel_blocks_are_complete_and_keep_pooled_n_clients():
    result = _available_global()
    models = models_and_win_rates_blocks(result)
    classifications = classifications_blocks(result)
    stability = portfolio_stability_blocks(result)
    events = portfolio_events_blocks(result)

    optimizer = next(frame for title, frame in models if title == "Modelos globales - SCP Classic Optimizer")
    core_optimizer = result.portfolio.model_tables.by_engine_block_model
    assert len(optimizer) == len(core_optimizer[core_optimizer["engine"] == "OPTIMIZER"])
    assert "Clientes con performance evaluable" in optimizer.columns
    assert optimizer["Clientes con performance evaluable"].max() == 2
    class_model = next(frame for title, frame in classifications if "clasificación del optimizer × modelo" in title.lower())
    assert len(class_model) == len(result.portfolio.optimizer.classification_model_tables)
    performance = next(frame for title, frame in stability if title == "Performance descriptiva por estabilidad")
    assert "Clientes con performance evaluable" in performance.columns
    event_table = next(frame for title, frame in events if "Eventos canónicos" in title)
    assert len(event_table) == len(result.portfolio.events.dataframe)


def test_global_excel_adds_18_19_and_preserves_native_formats_and_events(tmp_path: Path):
    result = _available_global()
    output = tmp_path / "global_portfolio.xlsx"
    build_global_workbook(result, output)
    workbook = openpyxl.load_workbook(output)

    assert workbook.sheetnames[-4:] == [
        "16_pareto_absolute_impact", "17_phase8_global",
        "18_portfolio_stability", "19_portfolio_events",
    ]
    assert workbook.sheetnames[11:13] == ["11_models_and_win_rates", "12_classifications"]
    model_sheet = workbook["11_models_and_win_rates"]
    ratio = _cell_below_header(model_sheet, "Cuota de selección sobre asignaciones posibles")
    improvement = _cell_below_header(
        model_sheet, "Mejora de WAPE — Optimizer vs Auto",
    )
    assert isinstance(ratio.value, (int, float)) and ratio.number_format == "0.0%"
    assert isinstance(improvement.value, (int, float)) and abs(improvement.value) > 1
    assert improvement.number_format == '+0.0"%";-0.0"%";0.0"%"'

    events = workbook["19_portfolio_events"]
    assert _cell_below_header(events, "WAPE — SCP Classic Auto").number_format == "0.0%"
    assert _cell_below_header(events, "Bias — SCP Classic Auto").number_format == "+0.0%;-0.0%;0.0%"
    event_header = next(
        cell for row in events.iter_rows() for cell in row
        if cell.value == "ID batch"
    )
    event_rows = [
        row for row in events.iter_rows(min_row=event_header.row + 1, values_only=True)
        if row[0] is not None
    ]
    assert len(event_rows) == len(result.portfolio.events.dataframe)


def test_global_excel_unavailable_and_empty_keep_stable_sheet_structure(tmp_path: Path):
    unavailable_path = tmp_path / "unavailable.xlsx"
    build_global_workbook(build_global_analysis_result(), unavailable_path)
    unavailable_book = openpyxl.load_workbook(unavailable_path)
    for name in (
        "11_models_and_win_rates", "12_classifications",
        "18_portfolio_stability", "19_portfolio_events",
    ):
        text = _sheet_text(unavailable_book[name])
        assert "selección por bloques no disponible" in text
        assert "AutoETS" not in text

    empty_path = tmp_path / "empty.xlsx"
    build_global_workbook(_empty_global(), empty_path)
    empty_book = openpyxl.load_workbook(empty_path)
    assert "no se observaron asignaciones" in _sheet_text(empty_book["11_models_and_win_rates"])
    assert "Cobertura de selección" in _sheet_text(empty_book["11_models_and_win_rates"])


def test_global_canonical_events_preserve_every_client_event_without_deduplication():
    result = _available_global()
    global_events = result.portfolio.events.dataframe
    client_events = [client.portfolio.events.dataframe for client in result.client_results]
    assert len(global_events) == sum(len(events) for events in client_events)
    assert set(global_events["ID_CLIENT"]) == {97001, 97002}
    assert set(global_events["engine"]) == {"SCP_AUTO", "OPTIMIZER"}


def test_global_surfaces_do_not_recalculate_or_recombine_portfolio(monkeypatch, tmp_path: Path):
    result = _available_global()
    presentation = prepare_portfolio_presentation(result.portfolio)

    def _boom(*args, **kwargs):
        raise AssertionError("global presentation must not recalculate or recombine portfolio")

    monkeypatch.setattr("src.global_analysis.combine_portfolio_analyses", _boom)
    for name in (
        "build_portfolio_analysis", "build_portfolio_event_table", "build_portfolio_coverage",
        "build_portfolio_model_result", "build_portfolio_optimizer_analysis",
        "build_portfolio_stability_analysis",
    ):
        module = __import__("src.portfolio", fromlist=[name])
        if hasattr(module, name):
            monkeypatch.setattr(f"src.portfolio.{name}", _boom)

    before = result.portfolio.events.dataframe.copy(deep=True)
    build_global_portfolio_view(presentation)
    build_global_report(result)
    build_global_workbook(result, tmp_path / "no_recalc.xlsx")
    assert_frame_equal(result.portfolio.events.dataframe, before)


def test_client_portfolio_helper_behavior_remains_identical_after_neutral_reuse():
    from src.client_portfolio_view import build_client_portfolio_view

    client = _client_result(_state_dataframe(97401))
    first = build_client_portfolio_view(client.portfolio)
    second = build_client_portfolio_view(client.portfolio)
    assert first == second
    assert first["models"][0]["rows"][0]["selection_count"] == "2"
