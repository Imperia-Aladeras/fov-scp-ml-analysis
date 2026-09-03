"""
Fase 8D -- auditoria: comprueba que el mismo valor de nucleo
(GlobalAnalysisResult.periods["6M"].phase8) llega SIN DIVERGENCIA de
contenido (solo diferencias de formato/representacion) a Excel, Markdown,
HTML (view-model) y HTML RENDERIZADO, para una fila de VOLUME_BUCKET, una
fila de modelo/clasificacion y una fila de classification_volume_cross
(incluida una fila SERIES_CLASSIFICATION x NOT_ASSIGNABLE). No recalcula
nada: toma como referencia phase8.volume_table/model_tables/
classification_volume_cross ya calculados por el nucleo
(build_phase8_global_multi_client_analysis_result en tests/factories.py) y
compara contra lo que produce cada writer/view-model.

Contrato de columnas (ver src/phase8.py, src/global_analysis.py, confirmado
en HEAD): model_tables, classification_tables y classification_volume_cross
llevan n_clients; volume_table global NUNCA lo lleva -- se verifica su
AUSENCIA consistente en las cuatro representaciones, no se compara su valor
(NO APLICA).
"""

from datetime import datetime
from pathlib import Path

import analysis_fov_scp_ml as pipeline
import openpyxl
import pytest

from src import html_view_models as vm
from src.global_analysis import analyze_global
from src.global_excel_writer import build_global_workbook
from src.global_report_writer import build_global_report
from src.html_report import generate_html_report
from src.phase8_presentation import direction_label_es, sort_volume_table, volume_bucket_label_es
from src.report_writer import (
    _fmt_num,
    _fmt_pct_fraction,
    _fmt_pct_scaled,
    _fmt_signed_pct_fraction,
    _fmt_signed_pct_scaled,
)
from tests.factories import (
    build_phase8_global_missing_client_results,
    build_phase8_global_multi_client_analysis_result,
    build_phase8_global_multi_client_results,
    build_phase8_global_null_classification_analysis_result,
)


def _xlsx_block_rows(path: Path, sheet: str, header_marker: str) -> list[dict]:
    """Lee UN bloque de write_blocks: filas desde la fila de cabecera cuyo primer valor es `header_marker`."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    rows_iter = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, row in enumerate(rows_iter) if row and row[0] == header_marker)
    headers = rows_iter[header_idx]
    result = []
    for row in rows_iter[header_idx + 1:]:
        if row[0] is None:
            break
        result.append({h: v for h, v in zip(headers, row) if h is not None})
    return result


def _render_global_html(global_result, tmp_path: Path, run_name: str):
    run_config = pipeline.build_run_config(
        pipeline.build_arg_parser(tmp_path).parse_args([
            "--input-dir", str(tmp_path / "data"), "--output-root", str(tmp_path / "runs"), "--run-name", run_name,
        ]),
        tmp_path,
    )
    now = datetime.now().astimezone()
    generate_html_report(
        run_config=run_config, results=global_result.client_results, global_result=global_result,
        all_outputs={}, global_outputs=[], execution_records=[],
        started_at=now, finished_at=now, status="OK",
        git_commit=None, git_worktree_dirty=None,
    )
    return (run_config.run_dir_temp / "index.html").read_text(encoding="utf-8")


# --------------------------------------------------------------------------
# Fila de VOLUME_BUCKET (n_clients NO APLICA -- ausente en las 4 representaciones)
# --------------------------------------------------------------------------

def test_volume_bucket_row_consistent_across_excel_markdown_html_view_model_and_rendered_html(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    core_table = sort_volume_table(phase8.volume_table)
    core_row = core_table.iloc[0]  # RELATIVE_LOW tras sort_volume_table
    assert core_row["category"] == "RELATIVE_LOW"
    assert "n_clients" not in core_table.columns  # NO APLICA: contrato del nucleo

    # --- Excel ---
    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    xlsx_rows = _xlsx_block_rows(xlsx_path, "17_phase8_global", "VOLUME_BUCKET")
    xlsx_row = next(r for r in xlsx_rows if r.get("VOLUME_BUCKET") == "Bajo relativo")
    assert "n_clients" not in xlsx_row and "N_CLIENTES" not in xlsx_row
    assert xlsx_row["n_comparable"] == core_row["n_comparable"]
    assert xlsx_row["Tasa victoria Optimizer"] == pytest.approx(core_row["win_rate_ml_pct"])
    assert xlsx_row["WAPE Auto"] == pytest.approx(core_row["scp_wape_agg"])
    assert xlsx_row["WAPE Optimizer"] == pytest.approx(core_row["ml_wape_agg"])
    assert xlsx_row["Mejora agregada Optimizer vs Auto"] == pytest.approx(core_row["improvement_agg_pct"])
    assert xlsx_row["Mediana mejora Optimizer vs Auto"] == pytest.approx(core_row["median_improvement_pct"])
    assert xlsx_row["Reduccion absoluta Optimizer vs Auto"] == pytest.approx(core_row["abs_error_reduction"])
    assert xlsx_row["pct_of_history_volume"] == pytest.approx(core_row["pct_of_history_volume"])
    assert xlsx_row["Bias Auto"] == pytest.approx(core_row["scp_bias_agg"])
    assert xlsx_row["Bias Optimizer"] == pytest.approx(core_row["ml_bias_agg"])
    assert xlsx_row["Direccion Auto"] == direction_label_es(core_row["scp_direction"])
    assert xlsx_row["Direccion Optimizer"] == direction_label_es(core_row["ml_direction"])
    assert bool(xlsx_row["small_sample"]) == bool(core_row["small_sample"])

    # --- Markdown ---
    report = build_global_report(result)
    section_22 = report.split("## 22.")[1]
    lines = [ln for ln in section_22.splitlines() if ln.strip().startswith("|")]
    headers = [h.strip() for h in lines[0].strip("|").split("|")]
    assert "N clientes" not in headers  # NO APLICA para volume_table
    md_row = None
    for line in lines[2:]:
        cells = [c.strip() for c in line.strip("|").split("|")]
        if cells[0] == "Bajo relativo":
            md_row = dict(zip(headers, cells))
            break
    assert md_row is not None, "fila 'Bajo relativo' no encontrada en la seccion 22"
    assert md_row["N comparable"] == _fmt_num(core_row["n_comparable"])
    assert md_row["Tasa victoria Optimizer"] == _fmt_pct_scaled(core_row["win_rate_ml_pct"])
    assert md_row["WAPE Auto"] == _fmt_pct_fraction(core_row["scp_wape_agg"])
    assert md_row["WAPE Optimizer"] == _fmt_pct_fraction(core_row["ml_wape_agg"])
    assert md_row["Mejora agregada Optimizer vs Auto"] == _fmt_signed_pct_scaled(core_row["improvement_agg_pct"])
    assert md_row["Mediana mejora Optimizer vs Auto"] == _fmt_signed_pct_scaled(core_row["median_improvement_pct"])
    assert md_row["Reduccion absoluta Optimizer vs Auto"] == _fmt_num(core_row["abs_error_reduction"])
    assert md_row["% volumen historico"] == _fmt_pct_scaled(core_row["pct_of_history_volume"])
    assert md_row["Bias Auto"] == _fmt_signed_pct_fraction(core_row["scp_bias_agg"])
    assert md_row["Bias Optimizer"] == _fmt_signed_pct_fraction(core_row["ml_bias_agg"])
    assert md_row["Direccion Auto"] == direction_label_es(core_row["scp_direction"])
    assert md_row["Direccion Optimizer"] == direction_label_es(core_row["ml_direction"])
    assert md_row["Muestra pequena"] == ("si" if core_row["small_sample"] else "no")

    # --- HTML view-model ---
    phase8_global_vm = vm.build_phase8_global_vm(phase8)
    html_row = next(r for r in phase8_global_vm["volume"]["rows"] if r["bucket"] == "Bajo relativo")
    assert "n_clients" not in html_row
    assert html_row["n"] == _fmt_num(core_row["n_comparable"])
    assert html_row["tasa_victoria_ml"] == _fmt_pct_scaled(core_row["win_rate_ml_pct"])
    assert html_row["wape_scp"] == _fmt_pct_fraction(core_row["scp_wape_agg"])
    assert html_row["wape_ml"] == _fmt_pct_fraction(core_row["ml_wape_agg"])
    assert html_row["mejora_agregada"] == _fmt_signed_pct_scaled(core_row["improvement_agg_pct"])
    assert html_row["mediana_mejora"] == _fmt_signed_pct_scaled(core_row["median_improvement_pct"])
    assert html_row["abs_error_reduction"] == _fmt_num(core_row["abs_error_reduction"])
    assert html_row["pct_of_history_volume"] == _fmt_pct_scaled(core_row["pct_of_history_volume"])
    assert html_row["scp_bias"] == _fmt_signed_pct_fraction(core_row["scp_bias_agg"])
    assert html_row["ml_bias"] == _fmt_signed_pct_fraction(core_row["ml_bias_agg"])
    assert html_row["scp_direction"] == direction_label_es(core_row["scp_direction"])
    assert html_row["ml_direction"] == direction_label_es(core_row["ml_direction"])
    assert html_row["muestra_pequena"] == bool(core_row["small_sample"])

    # --- HTML renderizado final (no solo el view-model) ---
    html = _render_global_html(result, tmp_path, "phase8_global_consistency_run")
    section = html.split('id="fase8-global"')[1].split("</section>")[0]
    assert html_row["tasa_victoria_ml"] in section
    assert html_row["wape_scp"] in section
    assert html_row["scp_bias"] in section
    assert html_row["ml_bias"] in section


# --------------------------------------------------------------------------
# Fila de modelo (n_clients SI aplica)
# --------------------------------------------------------------------------

def test_legacy_model_table_is_suppressed_across_excel_markdown_and_rendered_html(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    core_table = phase8.model_tables["ML_BEST_MODEL"]
    assert "AutoETS" in set(core_table["category"])  # payload interno de compatibilidad

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    excel_text = "\n".join(
        str(cell) for row in wb["11_models_and_win_rates"].iter_rows(values_only=True)
        for cell in row if cell is not None
    )
    assert "selección por bloques no disponible" in excel_text
    assert "ML_BEST_MODEL_OLDER_3M" in excel_text and "ML_BEST_MODEL_RECENT_3M" in excel_text
    assert "AutoETS" not in excel_text

    report = build_global_report(result)
    section_16 = report.split("## 16.")[1].split("## 17.")[0]
    assert "selección por bloques no disponible" in section_16
    assert "ML_BEST_MODEL_OLDER_3M" in section_16 and "ML_BEST_MODEL_RECENT_3M" in section_16
    assert "AutoETS" not in section_16

    html = _render_global_html(result, tmp_path, "phase8_global_model_consistency_run")
    section = html.split('id="portfolio-global"')[1].split("</section>")[0]
    assert "selección por bloques no disponible" in section
    assert "ML_BEST_MODEL_OLDER_3M" in section and "ML_BEST_MODEL_RECENT_3M" in section
    assert "AutoETS" not in section


# --------------------------------------------------------------------------
# Fila de classification_volume_cross (n_clients SI aplica), incluida la
# fila SERIES_CLASSIFICATION x NOT_ASSIGNABLE.
# --------------------------------------------------------------------------

def test_legacy_cross_table_is_suppressed_while_not_assignable_remains_visible(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    core_table = phase8.classification_volume_cross
    assert "lumpy" in set(core_table["SERIES_CLASSIFICATION"])  # payload interno de compatibilidad

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    excel_rows = list(wb["17_phase8_global"].iter_rows(values_only=True))
    assert not any(row and row[0] == "SERIES_CLASSIFICATION" for row in excel_rows)
    excel_text = "\n".join(str(cell) for row in excel_rows for cell in row if cell is not None)
    assert "Cruce no disponible" in excel_text
    assert "N_CLIENTES_VOLUMEN_NOT_ASSIGNABLE" in excel_text

    report = build_global_report(result)
    section_22 = report.split("## 22.")[1]
    assert "metadata legacy" in section_22
    assert "lumpy" not in section_22
    assert "Clientes con volumen relativo no asignable" in section_22

    html = _render_global_html(result, tmp_path, "phase8_global_cross_consistency_run")
    section = html.split('id="fase8-global"')[1].split("</section>")[0]
    assert "metadata legacy" in section
    assert "lumpy" not in section
    assert "Clientes con volumen relativo no asignable" in section


# --------------------------------------------------------------------------
# n_clients_with_not_assignable_volume: verbatim en las 4 representaciones
# --------------------------------------------------------------------------

def test_n_clients_with_not_assignable_volume_consistent_everywhere(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    expected = phase8.n_clients_with_not_assignable_volume
    assert expected == 1

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    xlsx_rows = _xlsx_block_rows(xlsx_path, "17_phase8_global", "N_CLIENTES_VOLUMEN_NOT_ASSIGNABLE")
    assert xlsx_rows[0]["N_CLIENTES_VOLUMEN_NOT_ASSIGNABLE"] == expected

    report = build_global_report(result)
    section_22 = report.split("## 22.")[1]
    assert str(expected) in section_22.split("no asignable")[1][:200]

    phase8_global_vm = vm.build_phase8_global_vm(phase8)
    assert phase8_global_vm["n_clients_not_assignable"] == _fmt_num(expected)

    html = _render_global_html(result, tmp_path, "phase8_global_notassignable_run")
    section = html.split('id="fase8-global"')[1].split("</section>")[0]
    assert _fmt_num(expected) in section


# --------------------------------------------------------------------------
# phase8 is None (compatibilidad, no reconstruir, no romper writers)
# --------------------------------------------------------------------------

def test_phase8_none_short_note_everywhere_no_fabricated_tables(tmp_path: Path):
    result = analyze_global(build_phase8_global_missing_client_results())
    assert result.periods["6M"].phase8 is None

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)  # no debe romper el workbook
    wb = openpyxl.load_workbook(xlsx_path)
    assert "17_phase8_global" in wb.sheetnames

    report = build_global_report(result)
    assert "## 22." in report
    assert "no disponible" in report.split("## 22.")[1]

    assert vm.build_phase8_global_vm(None) == {"available": False}

    html = _render_global_html(result, tmp_path, "phase8_global_none_consistency_run")
    section = html.split('id="fase8-global"')[1].split("</section>")[0]
    assert "no disponible" in section


# --------------------------------------------------------------------------
# Negativo: el cruce clasificacion x volumen NUNCA en outputs individuales
# ya existentes (8C, sin modificar). Si esto exigiera tocar un writer
# individual, el criterio del encargo es detenerse y reportarlo -- no se
# modifica ningun writer individual para que este test pase.
# --------------------------------------------------------------------------

def test_classification_volume_cross_absent_from_existing_individual_outputs(tmp_path: Path):
    from src.excel_writer import build_client_workbook
    from src.report_writer import build_client_report

    client_results = build_phase8_global_multi_client_results()
    client_result = client_results[0]

    md = build_client_report(client_result)
    assert "SERIES_CLASSIFICATION" not in md.split("## 18.")[1].split("## 19.")[0] or "VOLUME_BUCKET" not in md

    xlsx_path = tmp_path / "client.xlsx"
    build_client_workbook(client_result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["15_phase8_bias_volume"]
    header_rows = [row for row in ws.iter_rows(values_only=True) if row and "SERIES_CLASSIFICATION" in row]
    assert header_rows == []

    page_vm = vm.build_client_page_vm(client_result)
    assert "SERIES_CLASSIFICATION" not in str(page_vm.get("phase8", {}))


# --------------------------------------------------------------------------
# Paridad Markdown/Excel/HTML: las 2 perspectivas de modelo y las 4 de
# clasificacion deben poder consultarse EXPLICITAMENTE en las 3 salidas
# cuando phase8 esta disponible (cierre de Fase 8C/8D).
# --------------------------------------------------------------------------

def test_markdown_shows_explicit_notices_instead_of_legacy_model_and_classification_tables(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    report = build_global_report(result)

    section_16 = report.split("## 16.")[1].split("## 17.")[0]
    assert "selección por bloques no disponible" in section_16
    assert "AutoETS" not in section_16 and "x11 seasonal" not in section_16

    section_22 = report.split("## 22.")[1]
    assert "metadata legacy" in section_22
    for classification in ("smooth", "erratic", "intermittent", "lumpy"):
        assert classification not in section_22


def test_markdown_model_and_classification_tables_absent_without_bias_when_phase8_none(tmp_path: Path):
    result = analyze_global(build_phase8_global_missing_client_results())
    report = build_global_report(result)
    section_16 = report.split("## 16.")[1].split("## 17.")[0]
    assert "Modelos SCP" not in section_16  # comportamiento legacy exacto, sin fabricar nada nuevo


def test_html_and_excel_show_notices_instead_of_legacy_models_and_classifications(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws11 = wb["11_models_and_win_rates"]
    ws11_text = "\n".join(str(c) for row in ws11.iter_rows(values_only=True) for c in row if c is not None)
    assert "selección por bloques no disponible" in ws11_text
    assert "AutoETS" not in ws11_text and "x11 seasonal" not in ws11_text
    ws12 = wb["12_classifications"]
    ws12_text = "\n".join(str(c) for row in ws12.iter_rows(values_only=True) for c in row if c is not None)
    assert "selección por bloques no disponible" in ws12_text
    for category in ("smooth", "erratic", "intermittent", "lumpy"):
        assert category not in ws12_text

    html = _render_global_html(result, tmp_path, "phase8_global_parity_run")
    section = html.split('id="portfolio-global"')[1].split("</section>")[0]
    assert "selección por bloques no disponible" in section
    for category in ("AutoETS", "x11 seasonal", "smooth", "erratic", "intermittent", "lumpy"):
        assert category not in section


# --------------------------------------------------------------------------
# Clasificacion nula -> "(sin clasificar)" en las 4 representaciones,
# nunca "nan"/"None", nunca una fila/grupo aparte.
# --------------------------------------------------------------------------

def test_null_classification_is_normalized_internally_but_never_rendered(tmp_path: Path):
    result = build_phase8_global_null_classification_analysis_result()
    phase8 = result.periods["6M"].phase8
    assert phase8 is not None
    core_table = phase8.classification_tables["SERIES_CLASSIFICATION"]
    assert list(core_table["category"]) == ["smooth", "(sin clasificar)"]  # 1 unica fila, no duplicada
    assert core_table.loc[core_table["category"] == "(sin clasificar)", "n_comparable"].iloc[0] == 1

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws12_text = "\n".join(str(c) for row in wb["12_classifications"].iter_rows(values_only=True) for c in row if c is not None)
    assert "selección por bloques no disponible" in ws12_text
    assert "(sin clasificar)" not in ws12_text
    assert "nan" not in ws12_text.lower()
    assert "none" not in ws12_text.lower()

    report = build_global_report(result)
    section_22 = report.split("## 22.")[1]
    assert "metadata legacy" in section_22
    assert "(sin clasificar)" not in section_22
    assert "nan" not in section_22.lower()
    assert " none " not in f" {section_22.lower()} "

    html = _render_global_html(result, tmp_path, "phase8_global_null_class_run")
    section = html.split('id="fase8-global"')[1].split("</section>")[0]
    assert "metadata legacy" in section
    assert "(sin clasificar)" not in section
    assert "nan" not in section.lower()

    # El cruce global tambien normaliza la clasificacion nula (mismo MISSING_CATEGORY_LABEL).
    cross = phase8.classification_volume_cross
    assert "(sin clasificar)" in set(cross["SERIES_CLASSIFICATION"])


# --------------------------------------------------------------------------
# Truncado NUNCA silencioso del cruce clasificacion x volumen (Markdown/HTML).
# Excel conserva siempre la tabla completa (no se trunca ahi).
# --------------------------------------------------------------------------

def _inflate_cross_table_rows(phase8, n_target: int = 35):
    """
    Reutiliza el fixture multi-cliente real (todas las demas senales de
    phase8 siguen siendo las calculadas por el nucleo) y solo AMPLIA
    classification_volume_cross duplicando filas reales con SERIES_CLASSIFICATION
    distinta -- no fabrica columnas ni valores nuevos, solo repite filas ya
    validas hasta superar el top_n de presentacion (30), para poder probar
    el aviso de truncado sin construir 35 clientes reales.
    """
    import pandas as pd

    base = phase8.classification_volume_cross
    frames = []
    i = 0
    while sum(len(f) for f in frames) < n_target:
        copy = base.copy()
        copy["SERIES_CLASSIFICATION"] = copy["SERIES_CLASSIFICATION"].astype(str) + f"_dup{i}"
        frames.append(copy)
        i += 1
    inflated = pd.concat(frames, ignore_index=True).head(n_target)
    phase8.classification_volume_cross = inflated
    return inflated


def test_inflated_cross_table_remains_suppressed_in_markdown_html_and_excel(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    inflated = _inflate_cross_table_rows(phase8, n_target=35)
    assert len(inflated) == 35

    report = build_global_report(result)
    section_22 = report.split("## 22.")[1]
    assert "metadata legacy" in section_22
    assert "_dup" not in section_22

    html = _render_global_html(result, tmp_path, "phase8_global_truncation_run")
    section = html.split('id="fase8-global"')[1].split("</section>")[0]
    assert "metadata legacy" in section
    assert "_dup" not in section

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    excel_text = "\n".join(
        str(cell) for row in wb["17_phase8_global"].iter_rows(values_only=True)
        for cell in row if cell is not None
    )
    assert "Cruce no disponible" in excel_text
    assert "_dup" not in excel_text


def test_cross_table_no_truncation_note_when_all_rows_shown(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    assert len(phase8.classification_volume_cross) <= 30  # fixture real, sin inflar

    report = build_global_report(result)
    section_22 = report.split("## 22.")[1]
    assert "Mostrando" not in section_22.split("SERIES_CLASSIFICATION x VOLUME_BUCKET")[1]

    phase8_global_vm = vm.build_phase8_global_vm(phase8)
    assert phase8_global_vm["classification_volume_cross_total"] == len(phase8_global_vm["classification_volume_cross"])


def test_excel_cross_table_is_never_rendered_even_when_internal_payload_is_large(tmp_path: Path):
    result = build_phase8_global_multi_client_analysis_result()
    phase8 = result.periods["6M"].phase8
    inflated = _inflate_cross_table_rows(phase8, n_target=35)

    xlsx_path = tmp_path / "global.xlsx"
    build_global_workbook(result, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    rows = list(wb["17_phase8_global"].iter_rows(values_only=True))
    assert len(inflated) == 35
    assert not any(row and row[0] == "SERIES_CLASSIFICATION" for row in rows)
