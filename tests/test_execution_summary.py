from pathlib import Path

import openpyxl

from src.execution_summary import (
    INPUT_NOT_ANALYZED,
    build_execution_records,
    build_execution_summary_markdown,
    build_execution_summary_workbook,
    execution_summary_table,
)
from src.input_inventory import InputFileRecord
from tests.factories import build_multi_client_results


def _record_for(name: str, read_error: str | None = None) -> InputFileRecord:
    return InputFileRecord(
        name=name, relative_path=name, path=Path(name),
        size_bytes=None if read_error else 123, modified_at=None if read_error else "2026-01-01T00:00:00+00:00",
        mtime_ns=None if read_error else 1, sha256=None if read_error else "deadbeef" * 8,
        read_error=read_error,
    )


def _inventory_for_results(results) -> list[InputFileRecord]:
    return [_record_for(r.source.file_name) for r in results]


def test_build_execution_records_one_per_client():
    results = build_multi_client_results()
    inventory = _inventory_for_results(results)
    outputs_by_client = {r.source.id_client: [] for r in results}
    durations_by_client = {r.source.id_client: 0.42 for r in results}
    records = build_execution_records(inventory, results, outputs_by_client, durations_by_client)

    assert len(records) == 3
    assert {rec.id_client for rec in records} == {99999, 88888, 77777}
    assert all(rec.duracion_segundos == 0.42 for rec in records)
    assert all(rec.estado != INPUT_NOT_ANALYZED for rec in records)


def test_execution_summary_table_columns():
    results = build_multi_client_results()
    inventory = _inventory_for_results(results)
    outputs = {r.source.id_client: [f"outputs/{r.source.folder_name}/x.xlsx", f"outputs/{r.source.folder_name}/x.md",
                                     f"outputs/{r.source.folder_name}/processing_log_x.txt",
                                     "c.png", "c.png"] for r in results}
    durations = {r.source.id_client: 1.0 for r in results}
    records = build_execution_records(inventory, results, outputs, durations)
    table = execution_summary_table(records)

    assert list(table.columns) == [
        "ARCHIVO", "CARPETA_SALIDA", "ID_CLIENT", "ETIQUETA", "DISPLAY_NAME", "ID_BATCH", "ID_RUN_STAGING",
        "FILAS", "CANDIDATAS", "COMPARABLES_6M", "ESTADO", "WARNINGS", "ERRORS",
        "DURACION_SEGUNDOS", "INFORME_GENERADO", "EXCEL_GENERADO", "GRAFICOS_GENERADOS", "LOG_GENERADO",
        "TAMANO_BYTES", "SHA256", "ERROR_LECTURA",
    ]
    assert (table["GRAFICOS_GENERADOS"] == 2).all()
    assert table["INFORME_GENERADO"].all()
    assert table["EXCEL_GENERADO"].all()
    assert table["LOG_GENERADO"].all()
    assert table["SHA256"].notna().all()


def test_build_execution_summary_workbook_creates_file(tmp_path: Path):
    results = build_multi_client_results()
    inventory = _inventory_for_results(results)
    records = build_execution_records(inventory, results, {}, {})
    out_path = tmp_path / "execution_summary.xlsx"
    build_execution_summary_workbook(records, out_path)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["execution_summary"]
    ws = wb["execution_summary"]
    assert ws.max_row == 4  # cabecera + 3 clientes


def test_build_execution_summary_markdown_mentions_all_clients():
    results = build_multi_client_results()
    inventory = _inventory_for_results(results)
    records = build_execution_records(inventory, results, {}, {})
    markdown = build_execution_summary_markdown(records)
    for name in ("TA_FOV_SCP_ML_99999_Synthetic.csv", "TA_FOV_SCP_ML_88888_NoComparable.csv", "TA_FOV_SCP_ML_77777_AllMlWins.csv"):
        assert name in markdown


# --------------------------------------------------------------------------
# Fase 3 (cierre de inconsistencia, segunda vuelta): "Clientes procesados"
# debe contar solo registros con id_client asignado, no filas fisicas del
# resumen (un CSV con read_error genera una fila sin cliente asociado).
# --------------------------------------------------------------------------

def test_build_execution_summary_markdown_counts_only_records_with_id_client():
    """Escenario A: 1 CSV fisico particionado en varios clientes -> se cuentan todos (ninguno es id_client=None)."""
    results = build_multi_client_results()
    inventory = _inventory_for_results(results)
    records = build_execution_records(inventory, results, {}, {})
    assert all(rec.id_client is not None for rec in records)

    markdown = build_execution_summary_markdown(records)
    assert "**Clientes procesados:** 3" in markdown


def test_build_execution_summary_markdown_excludes_read_error_record_without_client():
    """Escenario B: 1 CSV fisico con read_error, sin ClientAnalysisResult -> Clientes procesados: 0."""
    inventory = [_record_for("TA_FOV_SCP_ML_66666_Broken.csv", read_error="permiso denegado (simulado)")]
    records = build_execution_records(inventory, [], {}, {})

    assert len(records) == 1
    assert records[0].id_client is None
    assert records[0].estado == INPUT_NOT_ANALYZED

    markdown = build_execution_summary_markdown(records)
    assert "**Clientes procesados:** 0" in markdown
    assert "**Clientes procesados:** 1" not in markdown


# --------------------------------------------------------------------------
# Correccion (punto 3): una fila por cada elemento del inventario, no solo
# por cada ClientAnalysisResult.
# --------------------------------------------------------------------------

def test_build_execution_records_includes_row_for_csv_with_no_result_at_all():
    """Un CSV con read_error (nunca produjo ClientSource ni ClientAnalysisResult) debe generar una fila."""
    results = build_multi_client_results()
    inventory = _inventory_for_results(results) + [
        _record_for("TA_FOV_SCP_ML_55555_Unreadable.csv", read_error="permiso denegado (simulado)")
    ]

    records = build_execution_records(inventory, results, {}, {})

    assert len(records) == 4
    ghost = next(r for r in records if r.archivo == "TA_FOV_SCP_ML_55555_Unreadable.csv")
    assert ghost.estado == INPUT_NOT_ANALYZED
    assert ghost.carpeta_salida == ""
    assert ghost.id_client is None
    assert ghost.filas is None
    assert ghost.candidatas is None
    assert ghost.comparables_6m is None
    assert ghost.analysis_error == "permiso denegado (simulado)"
    assert ghost.size_bytes is None
    assert ghost.sha256 is None
    assert ghost.log_generado is False


def test_build_execution_records_invalid_result_gets_folder_when_a_real_output_was_generated():
    """
    Correccion (punto 4): CARPETA_SALIDA se deriva de los outputs REALMENTE
    generados, no de file_valid. Un cliente invalido que si genero su
    processing_log (como ocurre siempre en la practica: _generate_client_outputs
    lo escribe incluso para clientes invalidos) debe informar su carpeta.
    Nunca se inventan informe, Excel ni graficos para un cliente invalido.
    """
    results = build_multi_client_results()
    invalid_result = next(r for r in results if "NoComparable" in r.source.file_label)
    invalid_result.file_valid = False
    invalid_result.status = "ERROR"

    inventory = _inventory_for_results(results)
    folder = invalid_result.source.folder_name
    outputs = {invalid_result.source.id_client: [f"outputs/{folder}/processing_log_{folder}.txt"]}
    records = build_execution_records(inventory, results, outputs, {})

    row = next(r for r in records if r.archivo == invalid_result.source.file_name)
    assert row.estado == "ERROR"
    assert row.carpeta_salida == f"outputs/{folder}/"
    assert row.log_generado is True
    assert row.informe_generado is False
    assert row.excel_generado is False
    assert row.graficos_generados == 0
    assert row.id_client == invalid_result.source.id_client


def test_build_execution_records_result_with_no_outputs_at_all_has_no_client_folder():
    """
    Si no se genero NINGUN output para un cliente correlacionado (p.ej. una
    excepcion aislada antes de escribir nada), CARPETA_SALIDA queda vacia
    aunque exista un ClientAnalysisResult: no hay carpeta real que informar.
    """
    results = build_multi_client_results()
    inventory = _inventory_for_results(results)
    records = build_execution_records(inventory, results, {}, {})  # outputs_by_file vacio para todos

    for row in records:
        assert row.carpeta_salida == ""
        assert row.log_generado is False


def test_build_execution_records_includes_row_when_inventory_has_read_error_and_no_result(tmp_path: Path):
    """
    Reproduce el escenario end-to-end del punto 4: un CSV cuyo inventario
    contiene read_error nunca produce un ClientAnalysisResult (no se copia,
    no se parsea), y debe seguir apareciendo en execution_summary.
    """
    inventory = [_record_for("TA_FOV_SCP_ML_66666_Broken.csv", read_error="No se pudo leer: acceso denegado")]
    records = build_execution_records(inventory, [], {}, {})

    assert len(records) == 1
    row = records[0]
    assert row.archivo == "TA_FOV_SCP_ML_66666_Broken.csv"
    assert row.estado == INPUT_NOT_ANALYZED
    assert row.analysis_error == "No se pudo leer: acceso denegado"
    assert row.carpeta_salida == ""

    markdown = build_execution_summary_markdown(records)
    assert "TA_FOV_SCP_ML_66666_Broken.csv" in markdown
    assert INPUT_NOT_ANALYZED in markdown

    out_path = tmp_path / "execution_summary.xlsx"
    build_execution_summary_workbook(records, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["execution_summary"]
    assert ws.max_row == 2  # cabecera + 1 fila
