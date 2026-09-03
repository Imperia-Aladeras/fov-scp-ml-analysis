from pathlib import Path

import openpyxl

from src.excel_writer import (
    EXEC_SUMMARY_PERIODS,
    _column_number_format,
    _dict_to_df,
    build_client_workbook,
    classifications_blocks,
    data_quality_checks_table,
    executive_summary_table,
    models_and_win_rates_blocks,
    pareto_absolute_impact_blocks,
    phase8_bias_volume_blocks,
)
from src.quality_checks import QualityIssue, Severity
from tests.factories import build_synthetic_client_result, build_volume_bucket_client_result

EXPECTED_SHEETS = [
    "00_readme", "01_executive_summary", "02_coverage_status", "03_semester", "04_first_quarter",
    "05_second_quarter", "06_monthly_summary", "07_monthly_winners", "08_models_and_win_rates",
    "09_classifications", "10_exclusions", "11_top_absolute_impact", "12_top_percentage_changes",
    "13_data_quality_checks", "14_pareto_absolute_impact", "15_phase8_bias_volume",
    "16_portfolio_stability", "17_portfolio_events",
]


def test_dict_to_df_sorted_descending():
    df = _dict_to_df({"a": 1, "b": 5, "c": 3}, "K", "V")
    assert df["K"].tolist() == ["b", "c", "a"]


# --------------------------------------------------------------------------
# Fase 9C: columnas DESCRIPCION/PERIODO en 13_data_quality_checks. Tests de
# PRESENTACION: los QualityIssue se inyectan a mano, no se ejercitan los
# checks de src/quality_checks.py.
# --------------------------------------------------------------------------

def test_data_quality_checks_table_keeps_code_severity_scope_and_context():
    """Case K: el Excel individual conserva code/severity/scope/periodo/mensaje sin perderlos."""
    result = build_synthetic_client_result(with_data=True)
    result.quality.issues.append(QualityIssue(
        Severity.WARNING, "NEGATIVE_NONNEGATIVE_METRIC_VALUE",
        "1 filas con SCP_WAPE_M1 negativo (dominio matematico >=0 para WAPE).",
        scope="period", details={"file": "f", "period": "M1", "column": "SCP_WAPE_M1", "n_violations": 1},
    ))
    result.quality.issues.append(QualityIssue(
        Severity.WARNING, "INVALID_BINARY_FLAG_VALUE",
        "1 filas con HAS_ML_EXCLUDED fuera del dominio {0,1}: {2: 1}.",
        scope="file", details={"file": "f", "column": "HAS_ML_EXCLUDED", "unexpected_values": {2: 1}, "n_rows": 1},
    ))
    df = data_quality_checks_table(result)

    period_row = df[df["CODIGO"] == "NEGATIVE_NONNEGATIVE_METRIC_VALUE"].iloc[0]
    assert period_row["SEVERIDAD"] == "WARNING"
    assert period_row["AMBITO"] == "period"
    assert period_row["PERIODO"] == "M1"
    assert period_row["DESCRIPCION"] == "Valor negativo en métrica no negativa"
    assert "SCP_WAPE_M1" in period_row["MENSAJE"]

    file_row = df[df["CODIGO"] == "INVALID_BINARY_FLAG_VALUE"].iloc[0]
    assert file_row["AMBITO"] == "file"
    assert file_row["PERIODO"] == ""  # scope file-level: sin periodo, no "None"
    assert file_row["DESCRIPCION"] == "Valor inválido en indicador binario"


def test_data_quality_checks_table_unknown_code_has_blank_description_not_invented():
    """Case H: fallback -- un codigo preexistente/desconocido no recibe una traduccion inventada."""
    result = build_synthetic_client_result(with_data=True)
    result.quality.issues.append(QualityIssue(
        Severity.WARNING, "EXTREME_WAPE", "3 filas con SCP_WAPE_M1 > 500%.", scope="period", details={"period": "M1", "n_extreme": 3},
    ))
    df = data_quality_checks_table(result)
    row = df[df["CODIGO"] == "EXTREME_WAPE"].iloc[0]
    assert row["DESCRIPCION"] == ""


def test_dict_to_df_empty():
    df = _dict_to_df({}, "K", "V")
    assert df.empty


def test_column_number_format_hints():
    assert _column_number_format("WAPE_SCP") == "0.0%"
    assert _column_number_format("MEJORA_RELATIVA_PCT") == '0.0"%"'
    assert _column_number_format("HISTORICO_TOTAL") == "#,##0"
    assert _column_number_format("ID_CONFIGURATION") is None


def test_executive_summary_table_has_one_row_per_period():
    result = build_synthetic_client_result(with_data=True)
    table = executive_summary_table(result)
    assert table["PERIODO"].tolist() == EXEC_SUMMARY_PERIODS
    row_6m = table[table["PERIODO"] == "6M"].iloc[0]
    assert row_6m["SERIES_COMPARABLES"] == 2


def test_build_client_workbook_creates_all_sheets(tmp_path: Path):
    result = build_synthetic_client_result(with_data=True)
    out_path = tmp_path / "summary.xlsx"
    build_client_workbook(result, out_path)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == EXPECTED_SHEETS
    assert wb["04_first_quarter"]["A1"].value == "Cobertura - 3 meses recientes (M3–M1)"
    assert wb["05_second_quarter"]["A1"].value == "Cobertura - 3 meses anteriores (M6–M4)"


def test_build_client_workbook_no_comparable_series_does_not_crash(tmp_path: Path):
    """
    Item explicito de la Fase 3: un cliente sin series comparables sigue
    siendo un caso valido; el Excel se genera igualmente sin inventar datos.
    """
    result = build_synthetic_client_result(with_data=False)
    out_path = tmp_path / "summary_empty.xlsx"
    build_client_workbook(result, out_path)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == EXPECTED_SHEETS

    ws = wb["01_executive_summary"]
    row_values = [cell.value for cell in ws[3]]  # primera fila de datos (6M)
    assert row_values[3] == 0  # SERIES_COMPARABLES = 0


# --------------------------------------------------------------------------
# 14_pareto_absolute_impact
# --------------------------------------------------------------------------

def test_pareto_absolute_impact_blocks_basic_content():
    result = build_synthetic_client_result(with_data=True)
    blocks = pareto_absolute_impact_blocks(result)
    titles = [t for t, _ in blocks]
    assert any("mejora" in t for t in titles)
    assert any("deterioro" in t for t in titles)
    assert any("Resumen de concentracion" in t for t in titles)

    improvement_table = next(df for t, df in blocks if "mejora" in t)
    deterioration_table = next(df for t, df in blocks if "deterioro" in t)
    assert improvement_table["ID_CONFIGURATION"].tolist() == [1001]
    assert deterioration_table["ID_CONFIGURATION"].tolist() == [1002]

    summary_table = next(df for t, df in blocks if "Resumen de concentracion" in t)
    assert summary_table["N_TOTAL"].tolist() == [1, 1]
    assert summary_table["N_NO_EVALUABLES"].tolist() == [0, 0]


def test_pareto_absolute_impact_blocks_reads_precomputed_pareto_without_recomputing():
    """
    La hoja 14 nunca debe volver a llamar a pareto_absolute_impact ni a
    build_pareto_analysis: las tablas devueltas deben ser exactamente los
    mismos objetos DataFrame ya almacenados en PeriodResult.pareto (una
    recomputacion produciria un DataFrame distinto, aunque tuviera el mismo
    contenido).
    """
    result = build_synthetic_client_result(with_data=True)
    pr = result.periods["6M"]
    blocks = {title: df for title, df in pareto_absolute_impact_blocks(result)}

    improvement_title = next(t for t in blocks if "mejora" in t)
    deterioration_title = next(t for t in blocks if "deterioro" in t)
    assert blocks[improvement_title] is pr.pareto.improvement.table
    assert blocks[deterioration_title] is pr.pareto.deterioration.table


def test_pareto_absolute_impact_blocks_all_improvement_client_notes_empty_deterioration(tmp_path: Path):
    from tests.factories import build_multi_client_results

    all_ml_result = build_multi_client_results()[2]  # 77777_AllMlWins: ambas filas ganan ML en 6M
    blocks = pareto_absolute_impact_blocks(all_ml_result)

    deterioration_table = next(df for t, df in blocks if "deterioro" in t)
    assert deterioration_table.empty
    note_texts = [v for t, df in blocks if t == "Nota" for v in df[""].tolist()]
    assert any("Sin series con deterioro" in n for n in note_texts)

    out_path = tmp_path / "all_ml.xlsx"
    build_client_workbook(all_ml_result, out_path)
    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == EXPECTED_SHEETS


# --------------------------------------------------------------------------
# Fase 8C: Bias integrado en 08/09, hoja nueva 15_phase8_bias_volume.
# --------------------------------------------------------------------------

def test_models_and_win_rates_blocks_show_portfolio_unavailable_notice():
    result = build_synthetic_client_result(with_data=True)
    blocks = models_and_win_rates_blocks(result)
    assert blocks[0][0] == "Estado del analisis"
    notice = blocks[0][1][""].iloc[0]
    assert "Análisis de selección por bloques no disponible" in notice
    assert "no implica un error del pipeline" in notice
    assert any(title == "Metadata especifica ausente" for title, _ in blocks)


def test_models_and_win_rates_blocks_reads_phase8_without_recomputing(monkeypatch):
    result = build_synthetic_client_result(with_data=True)

    def _boom(*args, **kwargs):
        raise AssertionError("excel_writer no debe recalcular category_performance_table_with_bias")

    monkeypatch.setattr("src.phase8.category_performance_table_with_bias", _boom)
    blocks = models_and_win_rates_blocks(result)
    assert blocks[0][0] == "Estado del analisis"


def test_classifications_blocks_show_portfolio_unavailable_notice():
    result = build_synthetic_client_result(with_data=True)
    blocks = classifications_blocks(result)
    assert blocks[0][0] == "Estado del analisis"
    notice = blocks[0][1][""].iloc[0]
    assert "Análisis de selección por bloques no disponible" in notice
    assert any(title == "Metadata especifica ausente" for title, _ in blocks)


def test_phase8_bias_volume_blocks_not_assignable_case():
    """El fixture sintetico estandar solo tiene 2 filas comparables en 6M -> NOT_ASSIGNABLE (n<3)."""
    result = build_synthetic_client_result(with_data=True)
    blocks = phase8_bias_volume_blocks(result)
    titles = [t for t, _ in blocks]
    assert any("Bias agregado del cliente" in t for t in titles)
    assert any("Volumen relativo" in t for t in titles)

    notes_table = next(df for t, df in blocks if t == "Notas metodologicas - Fase 8")
    notes = notes_table[""].tolist()
    assert any("NO asignable" in n for n in notes)
    assert any("6M" in n for n in notes)
    assert any("muestra pequena" in n.lower() for n in notes)
    assert any("routing" in n.lower() or "causalidad" in n.lower() for n in notes)


def test_phase8_bias_volume_blocks_ok_case_has_three_buckets_in_business_order():
    result = build_volume_bucket_client_result()
    blocks = phase8_bias_volume_blocks(result)
    volume_table = next(df for t, df in blocks if t.startswith("Volumen relativo"))
    assert volume_table["VOLUME_BUCKET"].tolist() == ["Bajo relativo", "Medio relativo", "Alto relativo"]
    for col in ("scp_bias_agg", "ml_bias_agg"):
        assert col in volume_table.columns


def test_phase8_bias_volume_blocks_never_recomputes(monkeypatch):
    result = build_volume_bucket_client_result()

    def _boom(*args, **kwargs):
        raise AssertionError("excel_writer no debe recalcular Fase 8")

    monkeypatch.setattr("src.phase8.build_phase8_client_diagnostics", _boom)
    monkeypatch.setattr("src.phase8.bias_aggregate", _boom)
    monkeypatch.setattr("src.phase8.compute_volume_buckets", _boom)
    blocks = phase8_bias_volume_blocks(result)
    assert any("Bias agregado del cliente" in t for t, _ in blocks)


def test_phase8_bias_volume_blocks_no_individual_classification_volume_cross(monkeypatch):
    """8C no implementa el cruce SERIES_CLASSIFICATION x VOLUME_BUCKET (exclusivo de 8D global)."""
    result = build_volume_bucket_client_result()

    def _boom(*args, **kwargs):
        raise AssertionError("el reporting individual no debe usar classification_volume_cross_table")

    monkeypatch.setattr("src.phase8.classification_volume_cross_table", _boom)
    blocks = phase8_bias_volume_blocks(result)
    for _, df in blocks:
        assert "SERIES_CLASSIFICATION" not in df.columns


def test_phase8_bias_volume_blocks_none_phase8_does_not_crash():
    result = build_synthetic_client_result(with_data=True)
    result.periods["6M"].phase8 = None
    blocks = phase8_bias_volume_blocks(result)
    assert any("no disponible" in t.lower() or "no disponible" in str(df).lower() for t, df in blocks)


def test_build_client_workbook_no_cross_classification_volume_sheet(tmp_path: Path):
    result = build_volume_bucket_client_result()
    out_path = tmp_path / "volume_buckets.xlsx"
    build_client_workbook(result, out_path)
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == EXPECTED_SHEETS
    ws = wb["15_phase8_bias_volume"]
    all_values = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "SERIES_CLASSIFICATION" not in all_values


def test_build_client_workbook_phase8_none_does_not_crash(tmp_path: Path):
    result = build_synthetic_client_result(with_data=True)
    result.periods["6M"].phase8 = None
    out_path = tmp_path / "phase8_none.xlsx"
    build_client_workbook(result, out_path)
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == EXPECTED_SHEETS
