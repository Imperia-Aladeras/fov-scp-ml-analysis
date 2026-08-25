"""
Generacion del Excel global (18 pestanas, 00_readme..17_phase8_global).
Comparativa entre todos los clientes con fichero valido.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

from src.excel_writer import _dict_to_df, _stats_dict_to_df, autosize_columns, write_blocks
from src.global_analysis import GlobalAnalysisResult, global_category_performance_table
from src.periods import ALL_PERIODS, MONTHLY_PERIODS, visible_label
from src.phase8_presentation import (
    BIAS_METHODOLOGY_NOTE,
    PHASE8_NO_ROUTING_NOTE,
    PHASE8_ONLY_6M_NOTE,
    PHASE8_SMALL_SAMPLE_NOTE,
    VOLUME_METHODOLOGY_NOTE_GLOBAL,
    direction_label_es,
    has_bias_columns,
    sort_volume_table,
    volume_bucket_label_es,
)
from src.quality_presentation import issue_period, metric_audit_friendly_label

MODEL_CLASSIFICATION_PERIOD = "6M"
_SINGLE_TABLE_SHEETS = ("15_data_quality_checks",)


def _translate_bias_directions(table: pd.DataFrame) -> pd.DataFrame:
    """
    Traduce scp_direction/ml_direction (POSITIVE/NEGATIVE/ZERO/NOT_EVALUABLE)
    a copy en castellano SOLO en una copia de presentacion; nunca toca las
    tablas de GlobalPeriodResult.phase8 ni recalcula ninguna columna. Mismo
    patron que src.excel_writer._translate_bias_directions (per-client), no
    importado desde alli porque los writers globales no dependen de
    excel_writer.py.
    """
    if not has_bias_columns(table):
        return table
    table = table.copy()
    table["scp_direction"] = table["scp_direction"].map(direction_label_es)
    table["ml_direction"] = table["ml_direction"].map(direction_label_es)
    return table


# --------------------------------------------------------------------------
# 00_readme
# --------------------------------------------------------------------------

def readme_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    # display_name (ID_CLIENT): el catalogo no garantiza nombres unicos entre
    # ID_CLIENT distintos, asi que el ID se muestra siempre junto al nombre.
    clients_ok = [f"{r.source.display_name} ({r.source.id_client})" for r in result.client_results]
    clients_invalid = [f"{r.source.display_name} ({r.source.id_client})" for r in result.invalid_results]
    lines = [
        f"Fecha de generacion: {pd.Timestamp.now():%Y-%m-%d %H:%M}",
        f"Clientes incluidos en la comparativa global ({len(clients_ok)}): {clients_ok}",
    ]
    if clients_invalid:
        lines.append(f"Clientes EXCLUIDOS por fichero invalido ({len(clients_invalid)}): {clients_invalid}")
    lines += [
        "",
        "Este libro compara clientes entre si. Cada cliente tiene ademas su propio Excel individual",
        "en outputs/<CLIENTE>/ con el detalle completo de su analisis.",
        "",
        "Cuatro perspectivas, deliberadamente separadas (no se mezclan):",
        "  1. Impacto global ponderado: SCP_WAPE_GLOBAL = SUM(error_absoluto) / SUM(historico) sobre",
        "     TODAS las series comparables de TODOS los clientes juntos. Responde: cuanto error total",
        "     reduce ML sobre el volumen total analizado.",
        "  2. Mejora por cliente: cada cliente pesa IGUAL, no se pondera por numero de series. La media,",
        "     mediana, percentiles y porcentaje de clientes que mejoran usan como denominador UNICAMENTE",
        "     los clientes con mejora calculable (N_CLIENTES_EVALUABLES = N_CLIENTES_TOTAL menos los que",
        "     no tienen ninguna serie comparable en el periodo), nunca el total de clientes cargados.",
        "  3. Mejora por serie: estadistica de la mejora relativa de cada serie individual de todos los",
        "     clientes juntos (recalculada desde las filas comparables originales, no reconstruida desde",
        "     medianas por cliente). Responde: como se comporta una serie tipica.",
        "  4. Impacto absoluto: se separan siempre REDUCCION_POSITIVA_TOTAL (suma de clientes que reducen",
        "     error) y DETERIORO_TOTAL_ABSOLUTO (suma de clientes que lo aumentan). No se calcula ni se",
        "     presenta un porcentaje de contribucion sobre la reduccion neta (REDUCCION_NETA) cuando esta",
        "     es cero o negativa: el porcentaje de cada cliente se calcula dentro de su propio grupo",
        "     (PCT_OF_POSITIVE_REDUCTION o PCT_OF_TOTAL_DETERIORATION), nunca sobre el neto.",
        "",
        "No se afirma 'ML mejora de forma generalizada' basandose unicamente en la perspectiva 1: se",
        "contrastan siempre las 4 perspectivas (ver 07_global_period_summary y el informe Markdown).",
        "",
        "Los clientes sin ninguna serie comparable en un periodo SI se incluyen en la cobertura, en los",
        "chequeos de calidad, en las tablas por cliente (02-06) y en execution_summary; unicamente quedan",
        "fuera del CALCULO de medias/medianas/WAPE/winners/mejoras de ese periodo (perspectivas 1-4) por no",
        "tener performance calculable, nunca de forma silenciosa: siempre se indica cuantos son.",
        "",
        "Modelos y clasificaciones (pestanas 11 y 12): semestre completo (6M), agregando todas las",
        "series comparables de todos los clientes.",
        "",
        "Periodos (nombre tecnico -> etiqueta visible):",
        *[f"  {p} -> {visible_label(p)}" for p in ALL_PERIODS],
    ]
    return [("README", pd.DataFrame({"": lines}))]


# --------------------------------------------------------------------------
# 01_executive_summary
# --------------------------------------------------------------------------

def executive_summary_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    rows = []
    for period in ALL_PERIODS:
        gp = result.periods[period]
        stats = gp.client_improvement_stats
        rows.append({
            "PERIODO": period, "ETIQUETA": gp.label,
            "N_CLIENTES_TOTAL": stats.get("n_total"),
            "N_CLIENTES_EVALUABLES": stats.get("n_evaluable"),
            "N_CLIENTES_SIN_PERFORMANCE": stats.get("n_missing"),
            "N_CLIENTES_MEJORAN": stats.get("n_improved"),
            "N_CLIENTES_EMPEORAN": stats.get("n_worse"),
            "N_CLIENTES_EMPATE": stats.get("n_tie"),
            "SERIES_CANDIDATAS": gp.n_candidates_total, "SERIES_COMPARABLES": gp.n_comparable_total,
            "PCT_COMPARABLE": gp.pct_comparable_global,
            "HISTORICO_TOTAL": gp.history_sum, "WAPE_SCP": gp.scp_wape_global, "WAPE_ML": gp.ml_wape_global,
            "MEJORA_GLOBAL_PONDERADA_PCT": gp.global_improvement_pct,
            "REDUCCION_ABSOLUTA_TOTAL": gp.abs_error_reduction_total,
            "MEDIA_MEJORA_POR_CLIENTE_PCT": stats.get("mean"),
            "MEDIANA_MEJORA_POR_CLIENTE_PCT": stats.get("median"),
            "MEDIA_MEJORA_POR_SERIE_PCT": gp.series_improvement_stats.get("mean"),
            "MEDIANA_MEJORA_POR_SERIE_PCT": gp.series_improvement_stats.get("median"),
            "PCT_CLIENTES_MEJORAN_SOBRE_EVALUABLES": stats.get("pct_improved"),
            "PCT_SERIES_GANA_ML": gp.winner_counts.get("ML", {}).get("pct"),
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 02_client_coverage
# --------------------------------------------------------------------------

def client_coverage_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    rows = []
    for r in result.client_results:
        row = {
            "ID_CLIENT": r.source.id_client, "ETIQUETA": r.source.file_label,
            "DISPLAY_NAME": r.source.display_name, "N_CANDIDATAS": r.n_candidates,
        }
        for period in ALL_PERIODS:
            pr = r.periods.get(period)
            row[f"PCT_COMPARABLE_{period}"] = pr.pct_comparable if pr else float("nan")
        pr_6m = r.periods.get("6M")
        row["N_ML_EXCLUDED"] = pr_6m.n_ml_excluded if pr_6m else None
        row["PCT_ML_EXCLUDED"] = pr_6m.pct_ml_excluded if pr_6m else None
        rows.append(row)
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 06_monthly_by_client (tabla compacta, formato ancho)
# --------------------------------------------------------------------------

def monthly_by_client_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    rows = []
    for r in result.client_results:
        row = {"ID_CLIENT": r.source.id_client, "ETIQUETA": r.source.file_label, "DISPLAY_NAME": r.source.display_name}
        for month in MONTHLY_PERIODS:
            pr = r.periods.get(month)
            row[f"PCT_COMPARABLE_{month}"] = pr.pct_comparable if pr else float("nan")
            row[f"MEJORA_{month}_PCT"] = pr.wape.get("improvement_pct") if pr else float("nan")
        rows.append(row)
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 07_global_period_summary
# --------------------------------------------------------------------------

def global_period_summary_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    rows = []
    for period in ALL_PERIODS:
        gp = result.periods[period]
        stats = gp.client_improvement_stats
        rows.append({
            "PERIODO": period, "ETIQUETA": gp.label,
            "HISTORICO_TOTAL": gp.history_sum,
            "ERROR_ABSOLUTO_SCP": gp.scp_abs_error_sum, "ERROR_ABSOLUTO_ML": gp.ml_abs_error_sum,
            "REDUCCION_ABSOLUTA": gp.abs_error_reduction_total,
            "WAPE_SCP": gp.scp_wape_global, "WAPE_ML": gp.ml_wape_global,
            "MEJORA_GLOBAL_PONDERADA_PCT": gp.global_improvement_pct,
            "N_CLIENTES_TOTAL": stats.get("n_total"),
            "N_CLIENTES_EVALUABLES": stats.get("n_evaluable"),
            "N_CLIENTES_SIN_PERFORMANCE": stats.get("n_missing"),
            "MEDIA_MEJORA_POR_CLIENTE_PCT": stats.get("mean"),
            "MEDIANA_MEJORA_POR_CLIENTE_PCT": stats.get("median"),
            "DESVIACION_ENTRE_CLIENTES": stats.get("std"),
            "MEDIA_MEJORA_POR_SERIE_PCT": gp.series_improvement_stats.get("mean"),
            "MEDIANA_MEJORA_POR_SERIE_PCT": gp.series_improvement_stats.get("median"),
            "PCT_CLIENTES_MEJORA_ML_SOBRE_EVALUABLES": stats.get("pct_improved"),
            "PCT_SERIES_GANA_ML": gp.winner_counts.get("ML", {}).get("pct"),
            "PCT_COMPARABLE": gp.pct_comparable_global,
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 08_client_improvement_stats / 09_series_improvement_stats
# --------------------------------------------------------------------------

def client_improvement_stats_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    return _stats_dict_to_df({period: result.periods[period].client_improvement_stats for period in ALL_PERIODS})


def series_improvement_stats_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    return _stats_dict_to_df({period: result.periods[period].series_improvement_stats for period in ALL_PERIODS})


# --------------------------------------------------------------------------
# 10_winner_distribution
# --------------------------------------------------------------------------

def winner_distribution_table(result: GlobalAnalysisResult) -> pd.DataFrame:
    rows = []
    for period in ALL_PERIODS:
        wc = result.periods[period].winner_counts
        rows.append({
            "PERIODO": period,
            "N_ML": wc.get("ML", {}).get("n", 0), "PCT_ML": wc.get("ML", {}).get("pct"),
            "N_SCP": wc.get("SCP", {}).get("n", 0), "PCT_SCP": wc.get("SCP", {}).get("pct"),
            "N_TIE": wc.get("TIE", {}).get("n", 0), "PCT_TIE": wc.get("TIE", {}).get("pct"),
            "N_TOTAL": wc.get("_total", 0),
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 11_models_and_win_rates / 12_classifications
# --------------------------------------------------------------------------

def models_and_win_rates_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    # GlobalPeriodResult.phase8 (calculado una unica vez en global_analysis.py)
    # es la fuente preferida -- ya incluye Bias y n_clients. Si es None (algun
    # cliente sin backend COMPARISON_STATUS en 6M), se mantiene el
    # comportamiento anterior a 8D, sin Bias.
    phase8 = result.periods[MODEL_CLASSIFICATION_PERIOD].phase8
    label = visible_label(MODEL_CLASSIFICATION_PERIOD)
    if phase8 is not None:
        ml_models = _translate_bias_directions(phase8.model_tables.get("ML_BEST_MODEL", pd.DataFrame()))
        scp_models = _translate_bias_directions(phase8.model_tables.get("SCP_BEST_MODEL", pd.DataFrame()))
    else:
        ml_models = global_category_performance_table(result.client_results, MODEL_CLASSIFICATION_PERIOD, "ML_BEST_MODEL")
        scp_models = global_category_performance_table(result.client_results, MODEL_CLASSIFICATION_PERIOD, "SCP_BEST_MODEL")
    blocks = [
        (f"Modelos ML (ML_BEST_MODEL) - {label} - todos los clientes", ml_models),
        (f"Modelos SCP (SCP_BEST_MODEL) - {label} - todos los clientes", scp_models),
    ]
    if phase8 is not None:
        blocks.append(("Nota metodologica - Bias", pd.DataFrame({"": [BIAS_METHODOLOGY_NOTE]})))
    return blocks


def classifications_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    phase8 = result.periods[MODEL_CLASSIFICATION_PERIOD].phase8
    label = visible_label(MODEL_CLASSIFICATION_PERIOD)
    blocks = []
    for col in ("ML_CLASSIFICATION", "ML_TYPE", "SERIES_CLASSIFICATION", "SCP_CLASSIFICATION"):
        if phase8 is not None:
            table = _translate_bias_directions(phase8.classification_tables.get(col, pd.DataFrame()))
        else:
            table = global_category_performance_table(result.client_results, MODEL_CLASSIFICATION_PERIOD, col)
        blocks.append((f"{col} - {label} - todos los clientes", table))
    if phase8 is not None:
        blocks.append(("Nota metodologica - Bias", pd.DataFrame({"": [BIAS_METHODOLOGY_NOTE]})))
    return blocks


# --------------------------------------------------------------------------
# 13_absolute_impact
# --------------------------------------------------------------------------

def absolute_impact_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    """
    Nunca se calcula ni se presenta un porcentaje de contribucion sobre la
    reduccion neta (que puede ser cero o negativa): se separan siempre los
    clientes que reducen error de los que lo aumentan, cada uno con su
    propio denominador (ver global_analysis._client_reduction_and_deterioration_tables).
    """
    gp = result.periods[MODEL_CLASSIFICATION_PERIOD]
    label = visible_label(MODEL_CLASSIFICATION_PERIOD)
    totals_df = pd.DataFrame([gp.reduction_totals])

    concentration_rows = []
    if not gp.client_reduction_table.empty:
        top_reducers = gp.client_reduction_table.sort_values("PCT_OF_POSITIVE_REDUCTION", ascending=False)
        concentration_rows.append({
            "GRUPO": "Clientes que reducen error",
            "TOP_1_CLIENTE_PCT_DEL_GRUPO": top_reducers["PCT_OF_POSITIVE_REDUCTION"].iloc[0],
            "TOP_3_CLIENTES_PCT_DEL_GRUPO": top_reducers["PCT_OF_POSITIVE_REDUCTION"].head(3).sum(),
            "N_CLIENTES_EN_GRUPO": len(gp.client_reduction_table),
        })
    if not gp.client_deterioration_table.empty:
        top_worseners = gp.client_deterioration_table.sort_values("PCT_OF_TOTAL_DETERIORATION", ascending=False)
        concentration_rows.append({
            "GRUPO": "Clientes que aumentan error",
            "TOP_1_CLIENTE_PCT_DEL_GRUPO": top_worseners["PCT_OF_TOTAL_DETERIORATION"].iloc[0],
            "TOP_3_CLIENTES_PCT_DEL_GRUPO": top_worseners["PCT_OF_TOTAL_DETERIORATION"].head(3).sum(),
            "N_CLIENTES_EN_GRUPO": len(gp.client_deterioration_table),
        })

    return [
        (f"Totales - {label} (REDUCCION_NETA = REDUCCION_POSITIVA_TOTAL - DETERIORO_TOTAL_ABSOLUTO)", totals_df),
        (f"Clientes que reducen error (ABS_ERROR_REDUCTION > 0) - {label}", gp.client_reduction_table),
        (f"Clientes que aumentan error (ABS_ERROR_REDUCTION < 0) - {label}", gp.client_deterioration_table),
        ("Concentracion dentro de cada grupo (nunca sobre la reduccion neta)", pd.DataFrame(concentration_rows)),
    ]


# --------------------------------------------------------------------------
# 16_pareto_absolute_impact
#
# Lee GlobalPeriodResult.pareto_series / pareto_clients (ya calculados una
# unica vez en global_analysis.py): esta hoja NUNCA llama a
# global_pareto_series/global_pareto_clients ni a build_pareto_analysis. No
# modifica conceptualmente 13_absolute_impact (client_reduction_table/
# client_deterioration_table siguen exactamente igual, sin tocar).
# --------------------------------------------------------------------------

def _pareto_concentration_summary_rows(gp) -> pd.DataFrame:
    rows = []
    for label, group, n_no_evaluables in (
        ("Series - mejora (ABS_ERROR_REDUCTION > 0)", gp.pareto_series.improvement, gp.pareto_series.n_no_evaluables),
        ("Series - deterioro (ABS_ERROR_REDUCTION < 0)", gp.pareto_series.deterioration, gp.pareto_series.n_no_evaluables),
        ("Clientes - mejora (ABS_ERROR_REDUCTION > 0)", gp.pareto_clients.improvement, gp.pareto_clients.n_no_evaluables),
        ("Clientes - deterioro (ABS_ERROR_REDUCTION < 0)", gp.pareto_clients.deterioration, gp.pareto_clients.n_no_evaluables),
    ):
        s = group.summary
        rows.append({
            "GRUPO": label, "N_TOTAL": s.n_total,
            "N_FOR_50": s.n_for_50, "N_FOR_80": s.n_for_80, "N_FOR_90": s.n_for_90,
            "TOTAL_IMPACT": s.total_impact, "N_NO_EVALUABLES": n_no_evaluables,
        })
    return pd.DataFrame(rows)


def pareto_absolute_impact_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    gp = result.periods[MODEL_CLASSIFICATION_PERIOD]
    if gp.pareto_series is None or gp.pareto_clients is None:
        return [("Sin datos", pd.DataFrame())]
    label = visible_label(MODEL_CLASSIFICATION_PERIOD)

    blocks: list[tuple[str, pd.DataFrame]] = [
        (f"Pareto series - mejora - {label} - todos los clientes", gp.pareto_series.improvement.table),
        (f"Pareto series - deterioro - {label} - todos los clientes", gp.pareto_series.deterioration.table),
        (f"Pareto clientes - mejora - {label}", gp.pareto_clients.improvement.table),
        (f"Pareto clientes - deterioro - {label}", gp.pareto_clients.deterioration.table),
        (f"Resumen de concentracion - {label}", _pareto_concentration_summary_rows(gp)),
    ]

    notes = []
    if gp.pareto_series.improvement.table.empty:
        notes.append("Sin series con mejora en 6M en el conjunto de clientes.")
    if gp.pareto_series.deterioration.table.empty:
        notes.append("Sin series con deterioro en 6M en el conjunto de clientes.")
    if gp.pareto_clients.improvement.table.empty:
        notes.append("Sin clientes con mejora en 6M.")
    if gp.pareto_clients.deterioration.table.empty:
        notes.append("Sin clientes con deterioro en 6M.")
    if gp.pareto_series.n_no_evaluables:
        notes.append(
            f"{gp.pareto_series.n_no_evaluables} serie(s) comparable(s) en 6M no son evaluables para "
            "impacto absoluto y no participan en el Pareto de series."
        )
    if gp.pareto_clients.n_no_evaluables:
        notes.append(
            f"{gp.pareto_clients.n_no_evaluables} cliente(s) no son evaluables para impacto absoluto "
            "en 6M y no participan en el Pareto de clientes."
        )
    if notes:
        blocks.append(("Nota", pd.DataFrame({"": notes})))
    return blocks


# --------------------------------------------------------------------------
# 17_phase8_global
#
# Lee GlobalPeriodResult.phase8 (ya calculado una unica vez en
# global_analysis.py): esta hoja NUNCA llama a build_phase8_global_diagnostics/
# bias_aggregate/compute_volume_buckets/classification_volume_cross_table. Las
# tablas se escriben tal cual las entrega el nucleo (incluye filas
# SERIES_CLASSIFICATION x NOT_ASSIGNABLE en el cruce si el nucleo las
# produce): no se filtran ni se inventan filas. volume_table NO tiene columna
# n_clients en el nucleo -- nunca se fabrica aqui.
# --------------------------------------------------------------------------

def _bias_total_table_global(bias_total) -> pd.DataFrame:
    return pd.DataFrame([
        {"METODO": "SCP", "BIAS_AGREGADO": bias_total.scp_bias_agg, "DIRECCION": direction_label_es(bias_total.scp_direction)},
        {"METODO": "ML", "BIAS_AGREGADO": bias_total.ml_bias_agg, "DIRECCION": direction_label_es(bias_total.ml_direction)},
    ])


def phase8_global_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    phase8 = result.periods[MODEL_CLASSIFICATION_PERIOD].phase8
    if phase8 is None:
        return [("Nota", pd.DataFrame({
            "": ["Diagnostico global Fase 8 (Bias/volumen) no disponible: al menos un cliente participante "
                 "no tiene el backend 6M (COMPARISON_STATUS) calculado."]
        }))]

    label = visible_label(MODEL_CLASSIFICATION_PERIOD)
    blocks: list[tuple[str, pd.DataFrame]] = [
        (f"Bias agregado global - {label}", _bias_total_table_global(phase8.bias_total)),
        ("Nota metodologica - Bias", pd.DataFrame({"": [BIAS_METHODOLOGY_NOTE]})),
    ]

    volume_table = _translate_bias_directions(sort_volume_table(phase8.volume_table))
    if volume_table is not None and not volume_table.empty:
        volume_table = volume_table.copy()
        volume_table["category"] = volume_table["category"].map(volume_bucket_label_es)
        volume_table = volume_table.rename(columns={"category": "VOLUME_BUCKET"})
    blocks.append((f"Volumen relativo global (VOLUME_BUCKET) - {label}", volume_table))

    cross_table = _translate_bias_directions(phase8.classification_volume_cross)
    if cross_table is not None and not cross_table.empty:
        cross_table = cross_table.copy()
        cross_table["VOLUME_BUCKET"] = cross_table["VOLUME_BUCKET"].map(volume_bucket_label_es)
        # Reordena columnas SOLO por legibilidad (SERIES_CLASSIFICATION/VOLUME_BUCKET/n_clients
        # primero): mismos valores, mismas columnas, nunca se recalcula ni se descarta ninguna.
        front_cols = [c for c in ("SERIES_CLASSIFICATION", "VOLUME_BUCKET", "n_clients") if c in cross_table.columns]
        cross_table = cross_table[front_cols + [c for c in cross_table.columns if c not in front_cols]]
    blocks.append((
        f"SERIES_CLASSIFICATION x VOLUME_BUCKET - {label} - exclusivo global",
        cross_table,
    ))

    blocks.append((
        "Clientes con volumen NOT_ASSIGNABLE",
        pd.DataFrame({"N_CLIENTES_VOLUMEN_NOT_ASSIGNABLE": [phase8.n_clients_with_not_assignable_volume]}),
    ))

    notes = [VOLUME_METHODOLOGY_NOTE_GLOBAL, PHASE8_ONLY_6M_NOTE, PHASE8_SMALL_SAMPLE_NOTE, PHASE8_NO_ROUTING_NOTE]
    blocks.append(("Notas metodologicas - Fase 8 global", pd.DataFrame({"": notes})))
    return blocks


# --------------------------------------------------------------------------
# 14_exclusions
# --------------------------------------------------------------------------

def exclusions_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    status_totals: dict[str, int] = {}
    reason_totals: dict[str, int] = {}
    per_client_rows = []
    for r in result.client_results:
        for status, n in r.comparison_status_distribution.items():
            status_totals[status] = status_totals.get(status, 0) + n
        pr_6m = r.periods.get("6M")
        if pr_6m:
            for reason, n in pr_6m.ml_exclusion_reason_counts.items():
                reason_totals[reason] = reason_totals.get(reason, 0) + n
            per_client_rows.append({
                "ID_CLIENT": r.source.id_client, "ETIQUETA": r.source.file_label,
                "DISPLAY_NAME": r.source.display_name,
                "N_ML_EXCLUDED": pr_6m.n_ml_excluded, "PCT_ML_EXCLUDED": pr_6m.pct_ml_excluded,
            })
    return [
        ("Distribucion global de COMPARISON_STATUS (todos los clientes)", _dict_to_df(status_totals, "COMPARISON_STATUS", "N")),
        ("Motivo de exclusion ML global (ML_EXCLUSION_REASON)", _dict_to_df(reason_totals, "MOTIVO", "N")),
        ("Exclusiones ML por cliente", pd.DataFrame(per_client_rows)),
    ]


# --------------------------------------------------------------------------
# 15_data_quality_checks
# --------------------------------------------------------------------------

def data_quality_checks_blocks(result: GlobalAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    all_results = result.client_results + result.invalid_results
    # Clave (severidad, codigo, PERIODO): un mismo codigo en dos periodos
    # distintos (p.ej. METRIC_001 en M1 y METRIC_001 en 6M) es una incidencia
    # distinta y no debe sumarse en la misma fila de resumen (Fase 9C,
    # seccion 12). Periodo=None para chequeos file/client-level.
    summary_counts: dict[tuple[str, str, str | None], dict] = {}
    detail_rows = []
    for r in all_results:
        for issue in r.quality.issues:
            if issue.severity.value == "OK":
                continue
            period = issue_period(issue)
            key = (issue.severity.value, issue.code, period)
            entry = summary_counts.setdefault(key, {"n_occurrences": 0, "clients": set()})
            entry["n_occurrences"] += 1
            # Identidad por ID_CLIENT, nunca por file_label/display_name (Fase
            # 5, bug corregido): varios ClientAnalysisResult de un mismo CSV
            # fisico comparten file_label (calculado una vez por fichero,
            # antes de particionar por ID_CLIENT), y el catalogo de nombres
            # tampoco garantiza display_name unico entre IDs distintos. Usar
            # cualquiera de los dos aqui infra-cuenta N_CLIENTES_AFECTADOS.
            entry["clients"].add(r.source.id_client)
            detail_rows.append({
                "ID_CLIENT": r.source.id_client, "CLIENTE": r.source.display_name,
                "SEVERIDAD": issue.severity.value,
                "CODIGO": issue.code, "DESCRIPCION": metric_audit_friendly_label(issue.code) or "",
                "AMBITO": issue.scope, "PERIODO": period or "",
                "MENSAJE": issue.message,
            })
    summary_rows = [
        {
            "SEVERIDAD": sev, "CODIGO": code, "DESCRIPCION": metric_audit_friendly_label(code) or "",
            "PERIODO": period or "", "N_OCURRENCIAS": v["n_occurrences"], "N_CLIENTES_AFECTADOS": len(v["clients"]),
        }
        for (sev, code, period), v in sorted(summary_counts.items(), key=lambda kv: -kv[1]["n_occurrences"])
    ]
    return [
        ("Resumen de incidencias por codigo (todos los clientes)", pd.DataFrame(summary_rows)),
        ("Detalle completo de incidencias", pd.DataFrame(detail_rows)),
    ]


# --------------------------------------------------------------------------
# Orquestacion
# --------------------------------------------------------------------------

def build_global_workbook(result: GlobalAnalysisResult, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_blocks(writer, "00_readme", readme_blocks(result))
        write_blocks(writer, "01_executive_summary", [("Resumen ejecutivo global por periodo", executive_summary_table(result))])
        write_blocks(writer, "02_client_coverage", [("Cobertura por cliente y periodo", client_coverage_table(result))])
        write_blocks(writer, "03_semester_by_client", [("Semestre completo (6M) por cliente", result.client_period_tables["6M"])])
        write_blocks(writer, "04_first_quarter_by_client", [(f"{visible_label('RECENT_3M')} por cliente", result.client_period_tables["RECENT_3M"])])
        write_blocks(writer, "05_second_quarter_by_client", [(f"{visible_label('OLDER_3M')} por cliente", result.client_period_tables["OLDER_3M"])])
        write_blocks(writer, "06_monthly_by_client", [("Cobertura y mejora mensual (M1-M6) por cliente", monthly_by_client_table(result))])
        write_blocks(writer, "07_global_period_summary", [("Resumen global por periodo", global_period_summary_table(result))])
        write_blocks(writer, "08_client_improvement_stats", [("Estadistica de mejora ENTRE CLIENTES (peso igual) por periodo", client_improvement_stats_table(result))])
        write_blocks(writer, "09_series_improvement_stats", [("Estadistica de mejora POR SERIE (todos los clientes juntos) por periodo", series_improvement_stats_table(result))])
        write_blocks(writer, "10_winner_distribution", [("Distribucion global de ganadores por periodo", winner_distribution_table(result))])
        write_blocks(writer, "11_models_and_win_rates", models_and_win_rates_blocks(result))
        write_blocks(writer, "12_classifications", classifications_blocks(result))
        write_blocks(writer, "13_absolute_impact", absolute_impact_blocks(result))
        write_blocks(writer, "14_exclusions", exclusions_blocks(result))
        write_blocks(writer, "15_data_quality_checks", data_quality_checks_blocks(result))
        write_blocks(writer, "16_pareto_absolute_impact", pareto_absolute_impact_blocks(result))
        write_blocks(writer, "17_phase8_global", phase8_global_blocks(result))

        for sheet_name in writer.sheets:
            autosize_columns(writer, sheet_name)
        for sheet_name in _SINGLE_TABLE_SHEETS:
            ws = writer.sheets[sheet_name]
            if ws.max_row >= 2 and ws.max_column >= 1:
                ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
