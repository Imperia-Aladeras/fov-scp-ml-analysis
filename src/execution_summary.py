"""
Resumen de ejecucion: execution_summary.md y .xlsx.

Una fila por CADA elemento del inventario de CSV de entrada (src.input_inventory),
no solo por cada ClientAnalysisResult: un CSV con read_error, o que por
cualquier motivo no llego a producir un resultado (p.ej. un fallo global a
mitad de procesamiento), debe seguir apareciendo con su nombre, tamano,
SHA-256 (o el error de lectura) y un estado explicito, sin inventar metricas
ni carpeta de cliente.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

from src.client_analysis import ClientAnalysisResult
from src.excel_writer import HEADER_FILL, HEADER_FONT, autosize_columns

INPUT_NOT_ANALYZED = "INPUT_NOT_ANALYZED"


@dataclass
class ExecutionRecord:
    archivo: str
    carpeta_salida: str
    id_client: object
    etiqueta: object
    display_name: object
    id_batch: list
    id_run_staging: list
    filas: object
    candidatas: object
    comparables_6m: object
    estado: str
    warnings: object
    errors: object
    duracion_segundos: float
    informe_generado: bool
    excel_generado: bool
    graficos_generados: int
    log_generado: bool
    size_bytes: object
    sha256: object
    analysis_error: object


def build_execution_records(
    inventory: list,
    results: list[ClientAnalysisResult],
    outputs_by_client: dict[int, list[str]],
    durations_by_client: dict[int, float],
    clients_subdir: str = "outputs",
) -> list[ExecutionRecord]:
    """
    Combina el inventario de CSV descubiertos con los resultados de analisis,
    agrupados por nombre de fichero fisico: un CSV puede haber producido
    varios ClientAnalysisResult (particion por ID_CLIENT), y cada uno genera
    su propia fila. outputs_by_client/durations_by_client estan indexados por
    ID_CLIENT, no por fichero. Un CSV sin ningun resultado correlacionado
    (nunca llego a parsearse: read_error, o no procesado por un fallo global)
    genera una unica fila con estado INPUT_NOT_ANALYZED, sin metricas
    inventadas y sin carpeta de cliente.
    """
    by_filename: dict[str, list[ClientAnalysisResult]] = {}
    for r in results:
        by_filename.setdefault(r.source.file_name, []).append(r)

    records: list[ExecutionRecord] = []
    for record in inventory:
        client_results = by_filename.get(record.name, [])

        if client_results:
            for r in client_results:
                source = r.source
                outputs = outputs_by_client.get(source.id_client, [])
                duration = durations_by_client.get(source.id_client, 0.0)
                counts = r.quality.summary_counts()
                pr_6m = r.periods.get("6M")
                log_generado = any(p.endswith(".txt") for p in outputs)
                # la carpeta de cliente se informa si se genero ALGUN output real
                # (como minimo el processing_log, que se escribe siempre que el
                # cliente se procesa, incluso si el fichero es invalido); nunca
                # se deriva solo de file_valid, que no dice nada sobre si la
                # carpeta llego a crearse en disco.
                carpeta_salida = f"{clients_subdir}/{source.folder_name}/" if outputs else ""
                records.append(ExecutionRecord(
                    archivo=record.name,
                    carpeta_salida=carpeta_salida,
                    id_client=source.id_client, etiqueta=source.file_label,
                    display_name=source.display_name,
                    id_batch=source.id_batch, id_run_staging=source.id_run_staging,
                    filas=source.n_rows, candidatas=r.n_candidates,
                    comparables_6m=pr_6m.n_comparable if pr_6m else 0,
                    estado=r.status, warnings=counts.get("WARNING", 0), errors=counts.get("ERROR", 0),
                    duracion_segundos=duration,
                    informe_generado=any(p.endswith(".md") for p in outputs),
                    excel_generado=any(p.endswith(".xlsx") for p in outputs),
                    graficos_generados=sum(1 for p in outputs if p.endswith(".png")),
                    log_generado=log_generado,
                    size_bytes=record.size_bytes, sha256=record.sha256, analysis_error=None,
                ))
        else:
            records.append(ExecutionRecord(
                archivo=record.name, carpeta_salida="",
                id_client=None, etiqueta=None, display_name=None, id_batch=[], id_run_staging=[],
                filas=None, candidatas=None, comparables_6m=None,
                estado=INPUT_NOT_ANALYZED, warnings=None, errors=None,
                duracion_segundos=0.0,
                informe_generado=False, excel_generado=False, graficos_generados=0,
                log_generado=False,
                size_bytes=record.size_bytes, sha256=record.sha256,
                analysis_error=record.read_error,
            ))
    return records


def _fmt_optional(x) -> str:
    return "n/d" if x is None else str(x)


def execution_summary_table(records: list[ExecutionRecord]) -> pd.DataFrame:
    return pd.DataFrame([{
        "ARCHIVO": rec.archivo, "CARPETA_SALIDA": rec.carpeta_salida,
        "ID_CLIENT": rec.id_client, "ETIQUETA": rec.etiqueta, "DISPLAY_NAME": rec.display_name,
        "ID_BATCH": str(rec.id_batch), "ID_RUN_STAGING": str(rec.id_run_staging),
        "FILAS": rec.filas, "CANDIDATAS": rec.candidatas, "COMPARABLES_6M": rec.comparables_6m,
        "ESTADO": rec.estado, "WARNINGS": rec.warnings, "ERRORS": rec.errors,
        "DURACION_SEGUNDOS": round(rec.duracion_segundos, 3),
        "INFORME_GENERADO": rec.informe_generado, "EXCEL_GENERADO": rec.excel_generado,
        "GRAFICOS_GENERADOS": rec.graficos_generados, "LOG_GENERADO": rec.log_generado,
        "TAMANO_BYTES": rec.size_bytes, "SHA256": rec.sha256, "ERROR_LECTURA": rec.analysis_error,
    } for rec in records])


def build_execution_summary_workbook(records: list[ExecutionRecord], output_path: Path) -> None:
    table = execution_summary_table(records)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="execution_summary", index=False)
        ws = writer.sheets["execution_summary"]
        for col_idx in range(1, len(table.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws.freeze_panes = "A2"
        autosize_columns(writer, "execution_summary")
        if ws.max_row >= 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def build_execution_summary_markdown(records: list[ExecutionRecord]) -> str:
    lines: list[str] = []
    a = lines.append

    # len(records) es el numero de filas de este resumen: una por cada
    # ClientAnalysisResult (Fase 3: un unico CSV fisico puede particionarse
    # en varios clientes), mas una por cada CSV con read_error que nunca
    # llego a producir un cliente (id_client=None, estado=INPUT_NOT_ANALYZED).
    # "Clientes procesados" cuenta solo los registros que representan
    # realmente un cliente: nunca el numero de filas ni el de CSV fisicos.
    n_clients_processed = sum(1 for rec in records if rec.id_client is not None)
    status_counts: dict[str, int] = {}
    for rec in records:
        status_counts[rec.estado] = status_counts.get(rec.estado, 0) + 1
    total_duration = sum(rec.duracion_segundos for rec in records)

    a("# Resumen de ejecucion")
    a("")
    a(f"**Fecha:** {pd.Timestamp.now():%d/%m/%Y %H:%M}")
    a(f"**Clientes procesados:** {n_clients_processed}")
    a(f"**Duracion total:** {total_duration:.2f} s")
    a("")
    a("| Estado | N |")
    a("|---|---|")
    for status, n in sorted(status_counts.items()):
        a(f"| {status} | {n} |")
    a("")
    a(
        "| Archivo | Nombre | Carpeta salida | ID_CLIENT | Filas | Candidatas | Comparables 6M | Estado | "
        "Warnings | Errors | Duracion (s) | Informe | Excel | Graficos | Log | Tamano (bytes) | SHA256 | Error de lectura |"
    )
    a("|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|")
    for rec in records:
        a(
            f"| {rec.archivo} | {_fmt_optional(rec.display_name)} | {rec.carpeta_salida} | {_fmt_optional(rec.id_client)} | {_fmt_optional(rec.filas)} | "
            f"{_fmt_optional(rec.candidatas)} | {_fmt_optional(rec.comparables_6m)} | {rec.estado} | "
            f"{_fmt_optional(rec.warnings)} | {_fmt_optional(rec.errors)} | {rec.duracion_segundos:.2f} | "
            f"{'si' if rec.informe_generado else 'no'} | {'si' if rec.excel_generado else 'no'} | "
            f"{rec.graficos_generados} | {'si' if rec.log_generado else 'no'} | {_fmt_optional(rec.size_bytes)} | "
            f"{_fmt_optional(rec.sha256)} | {_fmt_optional(rec.analysis_error)} |"
        )
    a("")
    return "\n".join(lines)
