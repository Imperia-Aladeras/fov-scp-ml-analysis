"""
Generacion del Excel individual por cliente (16 pestanas, 00_readme..15_phase8_bias_volume).

Formato aplicado (ver docs/analysis_requirements.md, seccion "Formato Excel"):
freeze panes en la primera fila, encabezados en negrita con relleno discreto,
anchos de columna ajustados, formato numerico coherente por tipo de columna
(WAPE como fraccion -> porcentaje nativo; metricas ya expresadas en base 100
-> sufijo "%" sin re-escalar; volumenes/errores -> separador de miles),
autofiltro en las pestanas de tabla unica (11, 12, 13). No se crean hojas
redundantes ni se convierte el libro en un dashboard complejo.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.client_analysis import ClientAnalysisResult
from src.models import category_performance_table, top_absolute_impact, top_percentage_changes
from src.periods import ALL_PERIODS, period_columns, visible_label
from src.phase8 import NOT_ASSIGNABLE
from src.phase8_presentation import (
    BIAS_METHODOLOGY_NOTE,
    PHASE8_NO_ROUTING_NOTE,
    PHASE8_ONLY_6M_NOTE,
    PHASE8_SMALL_SAMPLE_NOTE,
    VOLUME_METHODOLOGY_NOTE,
    direction_label_es,
    has_bias_columns,
    sort_volume_table,
    volume_bucket_label_es,
    volume_not_assignable_reason_es,
)

TITLE_FONT = Font(bold=True, size=11, color="1F3864")
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")

EXEC_SUMMARY_PERIODS = ["6M", "RECENT_3M", "OLDER_3M", "M1", "M2", "M3", "M4", "M5", "M6"]
MODEL_CLASSIFICATION_PERIOD = "6M"


def _column_number_format(col_name: str) -> str | None:
    name = col_name.upper()
    if "WAPE" in name or "BIAS" in name:
        return "0.0%"
    if any(h in name for h in ("PCT", "_RATE", "IMPROVEMENT")):
        return '0.0"%"'
    if any(h in name for h in ("HISTORICO", "HISTORY", "ABS_ERROR", "REDUCCION", "REDUCTION", "VOLUMEN")):
        return "#,##0"
    return None


def write_blocks(writer: pd.ExcelWriter, sheet_name: str, blocks: list[tuple[str, pd.DataFrame]]) -> None:
    startrow = 0
    for title, block_df in blocks:
        pd.DataFrame({title: []}).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=True)
        ws = writer.sheets[sheet_name]
        ws.cell(row=startrow + 1, column=1).font = TITLE_FONT
        startrow += 1
        if block_df is not None and not block_df.empty:
            block_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
            header_row = startrow + 1
            for col_idx, col_name in enumerate(block_df.columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                number_format = _column_number_format(str(col_name))
                if number_format:
                    for row_idx in range(header_row + 1, header_row + 1 + len(block_df)):
                        ws.cell(row=row_idx, column=col_idx).number_format = number_format
            startrow += len(block_df) + 2
        else:
            pd.DataFrame({"": ["(sin datos: ver nota en la pestana o en 00_readme)"]}).to_excel(
                writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False
            )
            startrow += 2
    writer.sheets[sheet_name].freeze_panes = "A2"


def autosize_columns(writer: pd.ExcelWriter, sheet_name: str, max_width: int = 45) -> None:
    ws = writer.sheets[sheet_name]
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), max_width)


def _dict_to_df(d: dict, key_name: str, value_name: str) -> pd.DataFrame:
    if not d:
        return pd.DataFrame({key_name: [], value_name: []})
    items = sorted(d.items(), key=lambda kv: -kv[1])
    return pd.DataFrame(items, columns=[key_name, value_name])


def _winner_counts_to_df(winner_counts: dict) -> pd.DataFrame:
    rows = []
    for method in ("ML", "SCP", "TIE"):
        entry = winner_counts.get(method, {"n": 0, "pct": float("nan")})
        rows.append({"METODO": method, "N": entry.get("n", 0), "PCT": entry.get("pct")})
    return pd.DataFrame(rows)


def _stats_dict_to_df(named_stats: dict[str, dict]) -> pd.DataFrame:
    rows = []
    for label, stats in named_stats.items():
        row = {"GRUPO": label}
        row.update(stats)
        rows.append(row)
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 00_readme
# --------------------------------------------------------------------------

def readme_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    source = result.source
    no_comparable_anywhere = result.periods and all(pr.n_comparable == 0 for pr in result.periods.values())
    lines = [
        f"Fuente: {source.file_name}",
        f"Cliente: {source.display_name} | ID_CLIENT={source.id_client} | etiqueta_fichero={source.file_label}",
        f"Batch: {source.id_batch} | Run staging: {source.id_run_staging} | Source run: {source.source_run_id}",
        f"Fecha de generacion: {pd.Timestamp.now():%Y-%m-%d %H:%M}",
        f"CSV reparado en memoria (comillas dobladas): {source.read_repaired}",
        f"Estado global del cliente: {result.status} | Fichero valido: {result.file_valid}",
        "",
        "Grano: ID_BATCH + ID_RUN_STAGING + ID_CLIENT + SOURCE_RUN_ID + ID_CONFIGURATION.",
        "Universo de cobertura: todas las filas candidatas (HAS_BASE_CANDIDATE=1) en M1..M6/",
        "RECENT_3M/OLDER_3M. En 6M/global el universo canonico es COMPARISON_STATUS='COMPARABLE'.",
        "Universo de performance: en periodos parciales (M1..M6, RECENT_3M, OLDER_3M), mascara de",
        "comparabilidad ESPECIFICA de cada periodo. En 6M/global la poblacion es exclusivamente",
        "COMPARISON_STATUS='COMPARABLE' (fuente de verdad backend); la mascara local se conserva",
        "solo como auditoria/reconciliacion. Ver pestana 02_coverage_status para el detalle por periodo.",
        "",
        "Periodos (nombre tecnico -> etiqueta visible):",
        *[f"  {p} -> {visible_label(p)}" for p in ALL_PERIODS],
        "",
        "Formulas principales:",
        "  WAPE_GLOBAL = SUM(error_absoluto_total) / SUM(historico_total)  [ponderado; nunca promedio",
        "  simple de WAPE por serie]",
        "  ML_IMPROVEMENT_VS_SCP_PCT = (SCP_WAPE - ML_WAPE) / SCP_WAPE * 100",
        "  ABS_ERROR_REDUCTION = SCP_TOTAL_ABS_ERROR - ML_TOTAL_ABS_ERROR (positivo = ML reduce error)",
        "",
        "Regla de comparabilidad en periodos parciales (M1..M6, RECENT_3M, OLDER_3M): pertenece al",
        "universo candidato, historico > 0, y forecast/error/WAPE validos para SCP y ML en ese periodo.",
        "Los trimestres usan directamente las columnas agregadas TOTAL_* ya materializadas en el CSV:",
        "no se exige que cada mes individual sea comparable de forma aislada.",
        "",
        "Regla de comparabilidad en 6M/global: exclusivamente COMPARISON_STATUS='COMPARABLE' (sin",
        "combinar con HAS_BASE_CANDIDATE). Una fila comparable permanece en la poblacion aunque le",
        "falte un input de una metrica concreta: esa metrica queda no evaluable (NaN), con su propio",
        "chequeo de calidad, sin excluir la fila.",
        "",
        "Regla de winner: WINNER_METHOD_* (columna original del CSV) es siempre la fuente de verdad.",
        "La formula exacta de 'relativeDiff' usada para generarla (regla de negocio: TIE cuando",
        "relativeDiff < 0.0001, salvo ambos WAPE=0 que siempre es TIE) no esta documentada en este",
        "repositorio. No se reconstruye ni se inventa ese umbral: solo se audita el caso totalmente",
        "especificado de ambos WAPE=0.",
        "",
        "Modelos y clasificaciones (pestanas 08 y 09): se muestran para el semestre completo (6M),",
        "la ventana mas representativa del cliente. 'Veces seleccionado' se calcula sobre el universo",
        "comparable de 6M, no sobre el total de candidatas.",
        "",
        "Limitaciones:",
        "  No se afirma mejora generalizada de ML basandose unicamente en el WAPE global: revisar",
        "  tambien la mediana de mejora por serie y la frecuencia de victoria (pestanas 03-07).",
        "  Los chequeos de calidad (pestana 13) son WARNING salvo problemas que invalidan el fichero;",
        "  una incidencia localizada en un mes no invalida los demas periodos ni el cliente.",
    ]
    if no_comparable_anywhere:
        lines.append(
            "  Este cliente NO tiene ninguna serie comparable en ningun periodo (ver 02_coverage_status "
            "para el motivo). Las pestanas de performance (03-09, 11, 12) se muestran vacias por diseno: "
            "es un caso valido de cobertura/diagnostico, no un error, y no se inventan metricas."
        )
    return [("README", pd.DataFrame({"": lines}))]


# --------------------------------------------------------------------------
# 01_executive_summary
# --------------------------------------------------------------------------

def executive_summary_table(result: ClientAnalysisResult) -> pd.DataFrame:
    rows = []
    for period in EXEC_SUMMARY_PERIODS:
        pr = result.periods.get(period)
        if pr is None:
            continue
        rows.append({
            "PERIODO": period, "ETIQUETA": pr.label,
            "SERIES_CANDIDATAS": pr.n_candidates, "SERIES_COMPARABLES": pr.n_comparable,
            "PCT_COMPARABLE": pr.pct_comparable,
            "HISTORICO_TOTAL": pr.wape.get("history_sum"),
            "WAPE_SCP": pr.wape.get("scp_wape_global"),
            "WAPE_ML": pr.wape.get("ml_wape_global"),
            "MEJORA_RELATIVA_PCT": pr.wape.get("improvement_pct"),
            "REDUCCION_ABSOLUTA": pr.abs_error_reduction_total,
            "PCT_GANA_ML": pr.winner_counts.get("ML", {}).get("pct"),
            "PCT_GANA_SCP": pr.winner_counts.get("SCP", {}).get("pct"),
            "PCT_EMPATE": pr.winner_counts.get("TIE", {}).get("pct"),
            "MEDIA_MEJORA_POR_SERIE_PCT": pr.improvement_stats_all.get("mean"),
            "MEDIANA_MEJORA_POR_SERIE_PCT": pr.improvement_stats_all.get("median"),
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 02_coverage_status
# --------------------------------------------------------------------------

def coverage_status_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    blocks: list[tuple[str, pd.DataFrame]] = [
        ("Distribucion global de COMPARISON_STATUS (universo candidato)",
         _dict_to_df(result.comparison_status_distribution, "COMPARISON_STATUS", "N")),
    ]
    pr = result.periods.get("6M")
    if pr is not None:
        blocks.append(("Exclusiones ML (HAS_ML_EXCLUDED=1, invariante por periodo)", pd.DataFrame([{
            "N_EXCLUIDAS": pr.n_ml_excluded, "PCT_SOBRE_CANDIDATAS": pr.pct_ml_excluded,
        }])))
        blocks.append(("Motivo de exclusion ML (ML_EXCLUSION_REASON)",
                        _dict_to_df(pr.ml_exclusion_reason_counts, "MOTIVO", "N")))
        blocks.append(("Motivo de ausencia de forecast SCP en 6M (SCP_NO_OUTPUT_REASON)",
                        _dict_to_df(pr.scp_no_output_reason_counts, "MOTIVO", "N")))

    reason_rows, status_rows = [], []
    for period, period_result in result.periods.items():
        for reason, n in period_result.not_comparable_reason_counts.items():
            reason_rows.append({"PERIODO": period, "MOTIVO_DERIVADO": reason, "N": n})
        for status, n in period_result.comparison_status_counts_not_comparable.items():
            status_rows.append({"PERIODO": period, "COMPARISON_STATUS": status, "N": n})
    blocks.append((
        "Motivo DERIVADO de no comparabilidad por periodo (especifico del nucleo, no sustituye a COMPARISON_STATUS)",
        pd.DataFrame(reason_rows),
    ))
    blocks.append((
        "COMPARISON_STATUS original (verbatim) entre las filas no comparables, por periodo",
        pd.DataFrame(status_rows),
    ))
    return blocks


# --------------------------------------------------------------------------
# 03_semester / 04_first_quarter / 05_second_quarter
# --------------------------------------------------------------------------

def period_detail_blocks(result: ClientAnalysisResult, period: str) -> list[tuple[str, pd.DataFrame]]:
    pr = result.periods.get(period)
    if pr is None:
        return [("Sin datos para este periodo", pd.DataFrame())]

    coverage_df = pd.DataFrame([{
        "SERIES_CANDIDATAS": pr.n_candidates, "SERIES_COMPARABLES": pr.n_comparable,
        "PCT_COMPARABLE": pr.pct_comparable, "SERIES_NO_COMPARABLES": pr.n_not_comparable,
    }])
    wape_df = pd.DataFrame([{
        "HISTORICO_TOTAL": pr.wape.get("history_sum"), "WAPE_SCP": pr.wape.get("scp_wape_global"),
        "WAPE_ML": pr.wape.get("ml_wape_global"), "MEJORA_RELATIVA_PCT": pr.wape.get("improvement_pct"),
        "REDUCCION_ABSOLUTA_ERROR": pr.abs_error_reduction_total,
    }])
    blocks = [
        (f"Cobertura - {pr.label}", coverage_df),
        ("WAPE global ponderado y mejora", wape_df),
        ("Distribucion de ganadores", _winner_counts_to_df(pr.winner_counts)),
        ("Estadistica de mejora relativa por serie (ML_IMPROVEMENT_VS_SCP)", _stats_dict_to_df({
            "TODAS_COMPARABLES": pr.improvement_stats_all, "GANA_ML": pr.improvement_stats_ml_wins,
            "GANA_SCP": pr.improvement_stats_scp_wins, "EMPATE": pr.improvement_stats_tie,
        })),
    ]
    if pr.n_comparable == 0:
        blocks.append(("Nota", pd.DataFrame({
            "": ["Sin series comparables en este periodo: no se calculan metricas de performance (no se inventan)."]
        })))
    return blocks


# --------------------------------------------------------------------------
# 06_monthly_summary / 07_monthly_winners
# --------------------------------------------------------------------------

def monthly_summary_table(result: ClientAnalysisResult) -> pd.DataFrame:
    rows = []
    for month in [f"M{i}" for i in range(1, 7)]:
        pr = result.periods.get(month)
        if pr is None:
            continue
        rows.append({
            "MES": month,
            "SERIES_CANDIDATAS": pr.n_candidates, "SERIES_COMPARABLES": pr.n_comparable,
            "PCT_COMPARABLE": pr.pct_comparable,
            "WAPE_SCP": pr.wape.get("scp_wape_global"), "WAPE_ML": pr.wape.get("ml_wape_global"),
            "MEJORA_RELATIVA_PCT": pr.wape.get("improvement_pct"),
            "REDUCCION_ABSOLUTA": pr.abs_error_reduction_total,
            "MEDIA_MEJORA_POR_SERIE_PCT": pr.improvement_stats_all.get("mean"),
            "MEDIANA_MEJORA_POR_SERIE_PCT": pr.improvement_stats_all.get("median"),
        })
    return pd.DataFrame(rows)


def monthly_winners_table(result: ClientAnalysisResult) -> pd.DataFrame:
    rows = []
    for month in [f"M{i}" for i in range(1, 7)]:
        pr = result.periods.get(month)
        if pr is None:
            continue
        wc = pr.winner_counts
        rows.append({
            "MES": month, "N_COMPARABLES": pr.n_comparable,
            "N_GANA_ML": wc.get("ML", {}).get("n", 0), "PCT_GANA_ML": wc.get("ML", {}).get("pct"),
            "N_GANA_SCP": wc.get("SCP", {}).get("n", 0), "PCT_GANA_SCP": wc.get("SCP", {}).get("pct"),
            "N_EMPATE": wc.get("TIE", {}).get("n", 0), "PCT_EMPATE": wc.get("TIE", {}).get("pct"),
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 08_models_and_win_rates / 09_classifications
# --------------------------------------------------------------------------

def _period_df_and_mask(result: ClientAnalysisResult, period: str):
    df = result.source.dataframe
    pr = result.periods.get(period)
    if df is None or pr is None or pr.comparable_mask is None:
        return None, None, None
    return df, period_columns(period), pr.comparable_mask


def _translate_bias_directions(table: pd.DataFrame) -> pd.DataFrame:
    """
    Traduce scp_direction/ml_direction (POSITIVE/NEGATIVE/ZERO/NOT_EVALUABLE)
    a copy en castellano SOLO en una copia de presentacion; nunca toca las
    tablas de PeriodResult.phase8 ni recalcula ninguna columna.
    """
    if not has_bias_columns(table):
        return table
    table = table.copy()
    table["scp_direction"] = table["scp_direction"].map(direction_label_es)
    table["ml_direction"] = table["ml_direction"].map(direction_label_es)
    return table


def models_and_win_rates_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    df, pcols, mask = _period_df_and_mask(result, MODEL_CLASSIFICATION_PERIOD)
    if df is None:
        return [("Sin datos", pd.DataFrame())]
    pr = result.periods.get(MODEL_CLASSIFICATION_PERIOD)
    # PeriodResult.phase8 (calculado una unica vez en client_analysis.py) es
    # la fuente preferida -- ya incluye Bias. Si es None (edge case: 6M sin
    # backend COMPARISON_STATUS), se mantiene el comportamiento anterior a
    # 8C, sin Bias, en vez de fallar.
    if pr is not None and pr.phase8 is not None:
        ml_models = _translate_bias_directions(pr.phase8.model_tables.get("ML_BEST_MODEL", pd.DataFrame()))
        scp_models = _translate_bias_directions(pr.phase8.model_tables.get("SCP_BEST_MODEL", pd.DataFrame()))
    else:
        ml_models = category_performance_table(df, pcols, mask, "ML_BEST_MODEL")
        scp_models = category_performance_table(df, pcols, mask, "SCP_BEST_MODEL")
    blocks = [
        (f"Modelos ML (ML_BEST_MODEL) - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", ml_models),
        (f"Modelos SCP (SCP_BEST_MODEL) - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", scp_models),
    ]
    if ml_models.empty and scp_models.empty:
        blocks.append(("Nota", pd.DataFrame({
            "": ["Sin series comparables en 6M: no se calculan modelos ni tasas de victoria (no se inventan)."]
        })))
    elif pr is not None and pr.phase8 is not None:
        blocks.append(("Nota metodologica - Bias", pd.DataFrame({"": [BIAS_METHODOLOGY_NOTE]})))
    return blocks


def classifications_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    df, pcols, mask = _period_df_and_mask(result, MODEL_CLASSIFICATION_PERIOD)
    if df is None:
        return [("Sin datos", pd.DataFrame())]
    pr = result.periods.get(MODEL_CLASSIFICATION_PERIOD)
    has_phase8 = pr is not None and pr.phase8 is not None
    blocks = []
    any_data = False
    for col, label in (
        ("ML_CLASSIFICATION", "ML_CLASSIFICATION"),
        ("ML_TYPE", "ML_TYPE"),
        ("SERIES_CLASSIFICATION", "SERIES_CLASSIFICATION"),
        ("SCP_CLASSIFICATION", "SCP_CLASSIFICATION"),
    ):
        if has_phase8:
            table = _translate_bias_directions(pr.phase8.classification_tables.get(col, pd.DataFrame()))
        else:
            table = category_performance_table(df, pcols, mask, col)
        any_data = any_data or not table.empty
        blocks.append((f"{label} - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", table))
    if not any_data:
        blocks.append(("Nota", pd.DataFrame({
            "": ["Sin series comparables en 6M: no se calculan clasificaciones (no se inventan)."]
        })))
    elif has_phase8:
        blocks.append(("Nota metodologica - Bias", pd.DataFrame({"": [BIAS_METHODOLOGY_NOTE]})))
    return blocks


# --------------------------------------------------------------------------
# 10_exclusions
# --------------------------------------------------------------------------

def exclusions_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    pr = result.periods.get("6M")
    n_status_excluded = result.comparison_status_distribution.get("NOT_COMPARABLE_ML_EXCLUDED", 0)
    n_flag_excluded = pr.n_ml_excluded if pr is not None else 0
    reconciliation = pd.DataFrame([{
        "N_COMPARISON_STATUS_ML_EXCLUDED": n_status_excluded,
        "N_HAS_ML_EXCLUDED_1": n_flag_excluded,
        "DIFERENCIA": n_flag_excluded - n_status_excluded,
    }])
    reason_df = _dict_to_df(pr.ml_exclusion_reason_counts, "MOTIVO", "N") if pr is not None else pd.DataFrame()
    return [
        ("Distribucion de COMPARISON_STATUS (universo candidato)",
         _dict_to_df(result.comparison_status_distribution, "COMPARISON_STATUS", "N")),
        ("Reconciliacion: COMPARISON_STATUS='NOT_COMPARABLE_ML_EXCLUDED' vs HAS_ML_EXCLUDED=1", reconciliation),
        ("Motivo de exclusion ML (ML_EXCLUSION_REASON)", reason_df),
    ]


# --------------------------------------------------------------------------
# 11_top_absolute_impact / 12_top_percentage_changes
# --------------------------------------------------------------------------

def top_absolute_impact_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    df, pcols, mask = _period_df_and_mask(result, MODEL_CLASSIFICATION_PERIOD)
    if df is None:
        return [("Sin datos", pd.DataFrame())]
    top_reduction, top_increase = top_absolute_impact(df, pcols, mask)
    return [
        (f"Top 20 mayor reduccion absoluta de error - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", top_reduction),
        (f"Top 20 mayor aumento absoluto de error - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", top_increase),
    ]


def top_percentage_changes_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    df, pcols, mask = _period_df_and_mask(result, MODEL_CLASSIFICATION_PERIOD)
    if df is None:
        return [("Sin datos", pd.DataFrame())]
    top_improve, top_worsen = top_percentage_changes(df, pcols, mask)
    return [
        (f"Top 20 mayor mejora porcentual - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", top_improve),
        (f"Top 20 mayor deterioro porcentual - {visible_label(MODEL_CLASSIFICATION_PERIOD)}", top_worsen),
    ]


# --------------------------------------------------------------------------
# 14_pareto_absolute_impact
#
# Lee PeriodResult.pareto (ya calculado una unica vez en client_analysis.py):
# esta hoja NUNCA llama a pareto_absolute_impact ni a build_pareto_analysis.
# Item separado de 11_top_absolute_impact (ranking crudo, sin porcentaje de
# contribucion): aqui se muestra el Pareto completo (RANK, PCT_OF_GROUP,
# CUMULATIVE_PCT) de cada grupo de signo, nunca mezclados.
# --------------------------------------------------------------------------

def _pareto_concentration_summary_table(pareto) -> pd.DataFrame:
    rows = []
    for label, group in (
        ("Mejora (ABS_ERROR_REDUCTION > 0)", pareto.improvement),
        ("Deterioro (ABS_ERROR_REDUCTION < 0)", pareto.deterioration),
    ):
        s = group.summary
        rows.append({
            "GRUPO": label, "N_TOTAL": s.n_total,
            "N_FOR_50": s.n_for_50, "N_FOR_80": s.n_for_80, "N_FOR_90": s.n_for_90,
            "TOTAL_IMPACT": s.total_impact,
            # Mismo valor en ambas filas: una serie no evaluable no tiene signo
            # conocido, no pertenece a un grupo concreto (ver src/pareto.py).
            "N_NO_EVALUABLES": pareto.n_no_evaluables,
        })
    return pd.DataFrame(rows)


def pareto_absolute_impact_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    pr = result.periods.get(MODEL_CLASSIFICATION_PERIOD)
    if pr is None or pr.pareto is None:
        return [("Sin datos", pd.DataFrame())]
    pareto = pr.pareto
    label = visible_label(MODEL_CLASSIFICATION_PERIOD)

    blocks: list[tuple[str, pd.DataFrame]] = [
        (f"Pareto de impacto absoluto - mejora (ABS_ERROR_REDUCTION > 0) - {label}", pareto.improvement.table),
        (f"Pareto de impacto absoluto - deterioro (ABS_ERROR_REDUCTION < 0) - {label}", pareto.deterioration.table),
        (f"Resumen de concentracion - {label}", _pareto_concentration_summary_table(pareto)),
    ]

    notes = []
    if pareto.improvement.table.empty:
        notes.append("Sin series con mejora (ABS_ERROR_REDUCTION > 0) en 6M para este cliente.")
    if pareto.deterioration.table.empty:
        notes.append("Sin series con deterioro (ABS_ERROR_REDUCTION < 0) en 6M para este cliente.")
    if pareto.n_no_evaluables:
        notes.append(
            f"{pareto.n_no_evaluables} serie(s) comparable(s) en 6M no son evaluables para impacto absoluto "
            "(falta SCP_TOTAL_ABS_ERROR_6M o ML_TOTAL_ABS_ERROR_6M): no participan en el Pareto."
        )
    if notes:
        blocks.append(("Nota", pd.DataFrame({"": notes})))
    return blocks


# --------------------------------------------------------------------------
# 13_data_quality_checks
# --------------------------------------------------------------------------

def data_quality_checks_table(result: ClientAnalysisResult) -> pd.DataFrame:
    rows = [
        {"SEVERIDAD": issue.severity.value, "CODIGO": issue.code, "AMBITO": issue.scope, "MENSAJE": issue.message}
        for issue in result.quality.issues
    ]
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 15_phase8_bias_volume
#
# Lee PeriodResult.phase8 (ya calculado una unica vez en client_analysis.py):
# esta hoja NUNCA llama a build_phase8_client_diagnostics/bias_aggregate/
# compute_volume_buckets. No existe cruce SERIES_CLASSIFICATION x
# VOLUME_BUCKET aqui (exclusivo de 8D global). Bias total y volumen son
# contenido genuinamente nuevo (no ampliaban ninguna hoja existente sin
# perder claridad), de ahi la hoja propia.
# --------------------------------------------------------------------------

def _bias_total_table(bias_total) -> pd.DataFrame:
    return pd.DataFrame([
        {"METODO": "SCP", "BIAS_AGREGADO": bias_total.scp_bias_agg, "DIRECCION": direction_label_es(bias_total.scp_direction)},
        {"METODO": "ML", "BIAS_AGREGADO": bias_total.ml_bias_agg, "DIRECCION": direction_label_es(bias_total.ml_direction)},
    ])


def phase8_bias_volume_blocks(result: ClientAnalysisResult) -> list[tuple[str, pd.DataFrame]]:
    pr = result.periods.get(MODEL_CLASSIFICATION_PERIOD)
    if pr is None or pr.phase8 is None:
        return [("Nota", pd.DataFrame({
            "": ["Fase 8 (Bias/volumen) no disponible: no se calculo el backend 6M (COMPARISON_STATUS) para este cliente."]
        }))]

    phase8 = pr.phase8
    label = visible_label(MODEL_CLASSIFICATION_PERIOD)
    blocks: list[tuple[str, pd.DataFrame]] = [
        (f"Bias agregado del cliente - {label}", _bias_total_table(phase8.bias_total)),
        ("Nota metodologica - Bias", pd.DataFrame({"": [BIAS_METHODOLOGY_NOTE]})),
    ]

    volume_table = _translate_bias_directions(sort_volume_table(phase8.volume_table))
    if volume_table is not None and not volume_table.empty:
        volume_table = volume_table.copy()
        volume_table["category"] = volume_table["category"].map(volume_bucket_label_es)
        volume_table = volume_table.rename(columns={"category": "VOLUME_BUCKET"})
    blocks.append((f"Volumen relativo (VOLUME_BUCKET) - {label}", volume_table))

    notes = [VOLUME_METHODOLOGY_NOTE, PHASE8_ONLY_6M_NOTE, PHASE8_SMALL_SAMPLE_NOTE, PHASE8_NO_ROUTING_NOTE]
    if phase8.volume.status == NOT_ASSIGNABLE:
        notes.insert(0, f"Volumen relativo NO asignable para este cliente: {volume_not_assignable_reason_es(phase8.volume.reason)}")
    blocks.append(("Notas metodologicas - Fase 8", pd.DataFrame({"": notes})))
    return blocks


# --------------------------------------------------------------------------
# Orquestacion
# --------------------------------------------------------------------------

# Autofiltro solo en la pestana de tabla verdaderamente unica: 11 y 12 tienen
# dos bloques apilados (top mejora / top deterioro) y un autofiltro sobre
# todo el rango mezclaria ambas tablas de forma confusa.
_SINGLE_TABLE_SHEETS = ("13_data_quality_checks",)


def build_client_workbook(result: ClientAnalysisResult, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_blocks(writer, "00_readme", readme_blocks(result))
        write_blocks(writer, "01_executive_summary", [("Resumen ejecutivo por periodo", executive_summary_table(result))])
        write_blocks(writer, "02_coverage_status", coverage_status_blocks(result))
        write_blocks(writer, "03_semester", period_detail_blocks(result, "6M"))
        write_blocks(writer, "04_first_quarter", period_detail_blocks(result, "RECENT_3M"))
        write_blocks(writer, "05_second_quarter", period_detail_blocks(result, "OLDER_3M"))
        write_blocks(writer, "06_monthly_summary", [("Resumen mensual M1-M6", monthly_summary_table(result))])
        write_blocks(writer, "07_monthly_winners", [("Ganadores por mes", monthly_winners_table(result))])
        write_blocks(writer, "08_models_and_win_rates", models_and_win_rates_blocks(result))
        write_blocks(writer, "09_classifications", classifications_blocks(result))
        write_blocks(writer, "10_exclusions", exclusions_blocks(result))
        write_blocks(writer, "11_top_absolute_impact", top_absolute_impact_blocks(result))
        write_blocks(writer, "12_top_percentage_changes", top_percentage_changes_blocks(result))
        write_blocks(writer, "13_data_quality_checks", [("Chequeos de calidad", data_quality_checks_table(result))])
        write_blocks(writer, "14_pareto_absolute_impact", pareto_absolute_impact_blocks(result))
        write_blocks(writer, "15_phase8_bias_volume", phase8_bias_volume_blocks(result))

        for sheet_name in writer.sheets:
            autosize_columns(writer, sheet_name)
        for sheet_name in _SINGLE_TABLE_SHEETS:
            ws = writer.sheets[sheet_name]
            if ws.max_row >= 2 and ws.max_column >= 1:
                ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
